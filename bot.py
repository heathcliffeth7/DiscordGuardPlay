import discord
from discord.ext import commands
import openpyxl
import os
import asyncio
from datetime import datetime, timedelta
from playwright.async_api import async_playwright
import difflib  # For fuzzy matching
from difflib import SequenceMatcher
import dotenv   # For .env file support
import re
import random
import string
import io
from collections import defaultdict, Counter
import time
import json
import threading
import signal
import copy
from pathlib import Path
import shlex
from typing import Awaitable, Callable, List, Optional, Set
_PIL_AVAILABLE = False
try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
    _PIL_AVAILABLE = True
    print("[PIL] Successfully imported PIL modules")
except ImportError as e:
    print(f"[PIL] Import failed - ImportError: {e}")
except Exception as e:
    print(f"[PIL] Import failed - Other error: {e}")

print(f"[PIL] Final status: _PIL_AVAILABLE = {_PIL_AVAILABLE}")
try:
    import regex as _advanced_regex_engine  # third-party 'regex' module (if available)
    _REGEX_ENGINE = _advanced_regex_engine
    _REGEX_ENGINE_NAME = "regex"
except Exception:  # pragma: no cover
    _REGEX_ENGINE = re
    _REGEX_ENGINE_NAME = "re"

# Load environment variables from .env file
dotenv.load_dotenv()

# Debug mode configuration
DEBUG_MODE = os.getenv("DEBUG_MODE", "false").lower() == "true"
if DEBUG_MODE:
    print("🐛 DEBUG MODE ENABLED - Sensitive information may be logged!")
else:
    print("🔒 PRODUCTION MODE - Debug logging disabled")

# Settings file path
SETTINGS_FILE = "bot_settings.json"
SECURITY_SETTINGS_FILE = "security_settings.json"

# Settings save/load functions
def save_settings():
    """Save all bot settings to JSON file"""
    settings = {
        "captcha_verify_role_id": captcha_verify_role_id,
        "captcha_panel_texts": captcha_panel_texts,
        "regex_settings_by_guild": {},
        "security_authorized_ids": list(security_authorized_ids),
        "play_authorized_ids": list(play_authorized_ids),
        "allowed_role_ids": allowed_role_ids,
        "no_avatar_filter_enabled": no_avatar_filter_enabled,
        "no_avatar_action": no_avatar_action,
        "no_avatar_timeout_duration": no_avatar_timeout_duration,
        "account_age_filter_enabled": account_age_filter_enabled,
        "account_age_min_days": account_age_min_days,
        "account_age_action": account_age_action,
        "account_age_timeout_duration": account_age_timeout_duration,
        "events": events,
        "event_nickname_limit": event_nickname_limit
    }
    
    # Convert regex settings to serializable format
    for guild_id, guild_rules in regex_settings_by_guild.items():
        settings["regex_settings_by_guild"][str(guild_id)] = {}
        for rule_name, rule_data in guild_rules.items():
            settings["regex_settings_by_guild"][str(guild_id)][rule_name] = {
                "pattern": rule_data.get("pattern", ""),
                "channels": list(rule_data.get("channels", set())),
                "exempt_users": list(rule_data.get("exempt_users", set())),
                "exempt_roles": list(rule_data.get("exempt_roles", set()))
            }
    
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=2, ensure_ascii=False)
        print(f"[SETTINGS] Settings saved to {SETTINGS_FILE}")
    except Exception as e:
        print(f"[SETTINGS] Error saving settings: {e}")

def load_settings():
    """Load all bot settings from JSON file"""
    global captcha_verify_role_id, captcha_panel_texts, regex_settings_by_guild
    global security_authorized_ids, play_authorized_ids, allowed_role_ids
    global no_avatar_filter_enabled, no_avatar_action, no_avatar_timeout_duration
    global account_age_filter_enabled, account_age_min_days, account_age_action, account_age_timeout_duration
    global events, event_nickname_limit
    
    if not os.path.exists(SETTINGS_FILE):
        print(f"[SETTINGS] Settings file {SETTINGS_FILE} not found, using defaults")
        return
    
    try:
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            settings = json.load(f)
        
        # Load basic settings
        captcha_verify_role_id = settings.get("captcha_verify_role_id")
        captcha_panel_texts.update(settings.get("captcha_panel_texts", {}))
        security_authorized_ids.update(settings.get("security_authorized_ids", []))
        play_authorized_ids.update(settings.get("play_authorized_ids", []))
        allowed_role_ids[:] = settings.get("allowed_role_ids", [])
        
        # Load security filter settings
        no_avatar_filter_enabled = settings.get("no_avatar_filter_enabled", False)
        no_avatar_action = settings.get("no_avatar_action")
        no_avatar_timeout_duration = settings.get("no_avatar_timeout_duration")
        account_age_filter_enabled = settings.get("account_age_filter_enabled", False)
        account_age_min_days = settings.get("account_age_min_days")
        account_age_action = settings.get("account_age_action")
        account_age_timeout_duration = settings.get("account_age_timeout_duration")
        
        # Load play events
        events.update(settings.get("events", {}))
        event_nickname_limit.update(settings.get("event_nickname_limit", {}))
        
        # Load regex settings and recompile patterns
        regex_data = settings.get("regex_settings_by_guild", {})
        for guild_id_str, guild_rules in regex_data.items():
            guild_id = int(guild_id_str)
            regex_settings_by_guild[guild_id] = {}
            for rule_name, rule_data in guild_rules.items():
                pattern = rule_data.get("pattern", "")
                if pattern:
                    try:
                        pattern_text, flags_letters = _parse_pattern_and_flags(pattern)
                        compiled = _compile_with_flags(pattern_text, flags_letters)
                        regex_settings_by_guild[guild_id][rule_name] = {
                            "pattern": pattern,
                            "compiled": compiled,
                            "channels": set(rule_data.get("channels", [])),
                            "exempt_users": set(rule_data.get("exempt_users", [])),
                            "exempt_roles": set(rule_data.get("exempt_roles", []))
                        }
                    except Exception as e:
                        print(f"[SETTINGS] Error recompiling regex pattern '{pattern}': {e}")
        
        print(f"[SETTINGS] Settings loaded successfully from {SETTINGS_FILE}")
        
    except Exception as e:
        print(f"[SETTINGS] Error loading settings: {e}")

# ============== SPAM STATISTICS PERSISTENCE ==============

def load_spam_violation_stats():
    """Load persisted spam violation statistics from disk."""
    global spam_violation_stats, spam_stats_loaded
    try:
        if SPAM_STATS_FILE.exists():
            with open(SPAM_STATS_FILE, "r", encoding="utf-8") as handle:
                spam_violation_stats = json.load(handle)
        else:
            spam_violation_stats = {}
    except Exception as exc:
        print(f"[SECURITY] Error loading spam stats: {exc}")
        spam_violation_stats = {}
    finally:
        spam_stats_loaded = True

def _parse_date_key(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return None

def _prune_spam_daily_counts(daily_counts):
    """Keep only the most recent configured number of days in daily counts."""
    if not daily_counts:
        return
    today = datetime.utcnow().date()
    cutoff = today - timedelta(days=MAX_SPAM_AGGREGATE_DAYS - 1)
    stale_keys = []
    for key in list(daily_counts.keys()):
        date_obj = _parse_date_key(key)
        if date_obj is None or date_obj < cutoff:
            stale_keys.append(key)
    for key in stale_keys:
        daily_counts.pop(key, None)

def _calculate_spam_aggregates(daily_counts):
    """Calculate window aggregates from per-day counts."""
    aggregates = {}
    today = datetime.utcnow().date()
    for label, days in SPAM_AGGREGATE_WINDOWS:
        cutoff = today - timedelta(days=days - 1)
        total = 0
        for key, value in daily_counts.items():
            date_obj = _parse_date_key(key)
            if date_obj is None:
                continue
            if date_obj >= cutoff:
                try:
                    total += int(value)
                except (TypeError, ValueError):
                    continue
        aggregates[label] = total
    return aggregates

async def _save_spam_violation_stats():
    """Persist spam violation statistics to disk."""
    async with spam_stats_lock:
        snapshot = copy.deepcopy(spam_violation_stats)

    def _write_snapshot():
        try:
            temp_path = SPAM_STATS_FILE.with_suffix(".tmp")
            with open(temp_path, "w", encoding="utf-8") as handle:
                json.dump(snapshot, handle, indent=2, ensure_ascii=False)
            temp_path.replace(SPAM_STATS_FILE)
        except Exception as exc:
            print(f"[SECURITY] Error saving spam stats: {exc}")

    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, _write_snapshot)

async def record_spam_violation(guild_id, user_id, rule_key, label=""):
    """Record a spam violation and update rolling aggregates."""
    global spam_stats_loaded
    if not spam_stats_loaded:
        load_spam_violation_stats()

    today_key = datetime.utcnow().strftime("%Y-%m-%d")

    async with spam_stats_lock:
        guild_key = str(guild_id)
        user_key = str(user_id)
        guild_bucket = spam_violation_stats.setdefault(guild_key, {})
        user_bucket = guild_bucket.setdefault(user_key, {})
        rule_bucket = user_bucket.setdefault(rule_key, {
            "label": label or rule_key,
            "daily_counts": {},
            "aggregates": {},
        })

        rule_bucket["label"] = label or rule_bucket.get("label") or rule_key

        daily_counts = rule_bucket.setdefault("daily_counts", {})
        daily_counts[today_key] = int(daily_counts.get(today_key, 0)) + 1

        _prune_spam_daily_counts(daily_counts)
        rule_bucket["aggregates"] = _calculate_spam_aggregates(daily_counts)
        rule_bucket["last_updated"] = today_key

    await _save_spam_violation_stats()

async def remove_spam_violation_stats_for_rule(guild_id, rule_key):
    """Remove stored violation statistics for a specific rule."""
    global spam_stats_loaded
    if not spam_stats_loaded:
        load_spam_violation_stats()

    guild_key = str(guild_id)
    async with spam_stats_lock:
        guild_bucket = spam_violation_stats.get(guild_key)
        if not guild_bucket:
            return

        empty_users = []
        for user_key, user_bucket in guild_bucket.items():
            if rule_key in user_bucket:
                user_bucket.pop(rule_key, None)
            if not user_bucket:
                empty_users.append(user_key)

        for user_key in empty_users:
            guild_bucket.pop(user_key, None)

        if not guild_bucket:
            spam_violation_stats.pop(guild_key, None)

    await _save_spam_violation_stats()

def _reset_spam_history_for_rule(guild_id: int, rule_key: str) -> None:
    """Reset cached spam counters so a rule restarts fresh."""
    keys_to_delete = [key for key in spam_message_history if key[0] == guild_id]
    for history_key in keys_to_delete:
        spam_message_history.pop(history_key, None)

    trigger_keys = [
        key for key in spam_rule_trigger_log
        if key[0] == guild_id and key[2] == rule_key
    ]
    for trigger_key in trigger_keys:
        spam_rule_trigger_log.pop(trigger_key, None)

# ============== SECURITY SETTINGS PERSISTENCE ==============

def save_security_settings():
    """Save all security settings to JSON file"""
    try:
        # Convert regex settings to serializable format
        serializable_regex_settings = {}
        for guild_id, guild_rules in regex_settings_by_guild.items():
            serializable_regex_settings[str(guild_id)] = {}
            for rule_name, rule_data in guild_rules.items():
                serializable_regex_settings[str(guild_id)][rule_name] = {
                    "pattern": rule_data.get("pattern", ""),
                    "channels": list(rule_data.get("channels", set())),
                    "exempt_users": list(rule_data.get("exempt_users", set())),
                    "exempt_roles": list(rule_data.get("exempt_roles", set()))
                }
        
        # Serialize spam rules
        serializable_spam_rules = {}
        for guild_id, guild_rules in spam_rules_by_guild.items():
            serializable_spam_rules[str(guild_id)] = {}
            for rule_name, rule_data in guild_rules.items():
                serializable_spam_rules[str(guild_id)][rule_name] = {
                    "label": rule_data.get("label", rule_name),
                    "min_length": rule_data.get("min_length", 0),
                    "similarity_threshold": rule_data.get("similarity_threshold", 0.0),
                    "time_window": rule_data.get("time_window", 0),
                    "message_count": rule_data.get("message_count", 0),
                    "dm_message": rule_data.get("dm_message", ""),
                    "notify_channel_id": rule_data.get("notify_channel_id"),
                    "channels": list(rule_data.get("channels", set())),
                    "excluded_channels": list(rule_data.get("excluded_channels", set())),
                    "targeted_roles": list(rule_data.get("targeted_roles", set())),
                    "exempted_roles": list(rule_data.get("exempted_roles", set())),
                    "nonreply_only": rule_data.get("nonreply_only", False),
                    "mod_action": rule_data.get("mod_action"),
                    "regex_pattern": rule_data.get("regex_pattern"),  # None for similarity mode
                }
        
        # Serialize captcha panel texts
        serializable_panel_texts = {}
        for guild_id, panel_data in captcha_panel_texts.items():
            serializable_panel_texts[str(guild_id)] = panel_data
        
        settings_data = {
            "version": "1.0",
            "timestamp": time.time(),
            
            # Global Security Filters
            "no_avatar_filter_enabled": no_avatar_filter_enabled,
            "no_avatar_action": no_avatar_action,
            "no_avatar_timeout_duration": no_avatar_timeout_duration,
            
            "account_age_filter_enabled": account_age_filter_enabled,
            "account_age_min_days": account_age_min_days,
            "account_age_action": account_age_action,
            "account_age_timeout_duration": account_age_timeout_duration,
            
            # Security Authorization
            "security_authorized_ids": list(security_authorized_ids),
            
            # Whitelist
            "security_whitelist_users": list(security_whitelist_users),
            
            # CAPTCHA Settings
            "captcha_verify_role_id": captcha_verify_role_id,
            "captcha_panel_texts": serializable_panel_texts,
            
            # Regex Settings
            "regex_settings_by_guild": serializable_regex_settings,
            
            # Spam Settings
            "spam_rules_by_guild": serializable_spam_rules,
            
            # Verify button usage (for statistics only)
            "verify_button_usage": dict(verify_button_usage)
        }
        
        # Backup existing file if present
        if os.path.exists(SECURITY_SETTINGS_FILE):
            backup_file = f"{SECURITY_SETTINGS_FILE}.backup"
            try:
                os.rename(SECURITY_SETTINGS_FILE, backup_file)
                print(f"[SECURITY] Backup created: {backup_file}")
            except Exception as e:
                print(f"[SECURITY] Warning: Could not create backup: {e}")
        
        # Save new settings
        with open(SECURITY_SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings_data, f, indent=2, ensure_ascii=False)
        
        print(f"[SECURITY] Settings saved successfully to {SECURITY_SETTINGS_FILE}")
        return True
        
    except Exception as e:
        print(f"[SECURITY] Error saving settings: {e}")
        return False

def load_security_settings():
    """Load security settings from JSON file"""
    global no_avatar_filter_enabled, no_avatar_action, no_avatar_timeout_duration
    global account_age_filter_enabled, account_age_min_days, account_age_action, account_age_timeout_duration
    global security_authorized_ids, security_whitelist_users, captcha_verify_role_id, captcha_panel_texts
    global regex_settings_by_guild, verify_button_usage, spam_rules_by_guild
    
    try:
        if not os.path.exists(SECURITY_SETTINGS_FILE):
            print(f"[SECURITY] No settings file found at {SECURITY_SETTINGS_FILE}, using defaults")
            return False
        
        with open(SECURITY_SETTINGS_FILE, 'r', encoding='utf-8') as f:
            settings_data = json.load(f)
        
        # Version control
        version = settings_data.get("version", "unknown")
        print(f"[SECURITY] Loading settings version: {version}")
        
        # Global Security Filters
        no_avatar_filter_enabled = settings_data.get("no_avatar_filter_enabled", False)
        no_avatar_action = settings_data.get("no_avatar_action", None)
        no_avatar_timeout_duration = settings_data.get("no_avatar_timeout_duration", None)
        
        account_age_filter_enabled = settings_data.get("account_age_filter_enabled", False)
        account_age_min_days = settings_data.get("account_age_min_days", None)
        account_age_action = settings_data.get("account_age_action", None)
        account_age_timeout_duration = settings_data.get("account_age_timeout_duration", None)
        
        # Security Authorization
        security_authorized_ids = set(settings_data.get("security_authorized_ids", []))
        
        # Whitelist
        security_whitelist_users = set(settings_data.get("security_whitelist_users", []))
        
        # CAPTCHA Settings
        captcha_verify_role_id = settings_data.get("captcha_verify_role_id", None)
        
        # Panel texts
        panel_texts_data = settings_data.get("captcha_panel_texts", {})
        captcha_panel_texts.clear()
        for guild_id_str, panel_data in panel_texts_data.items():
            try:
                guild_id = int(guild_id_str)
                captcha_panel_texts[guild_id] = panel_data
            except ValueError:
                print(f"[SECURITY] Warning: Invalid guild ID in panel texts: {guild_id_str}")
        
        # Regex Settings
        regex_data = settings_data.get("regex_settings_by_guild", {})
        regex_settings_by_guild.clear()
        for guild_id_str, guild_rules in regex_data.items():
            try:
                guild_id = int(guild_id_str)
                regex_settings_by_guild[guild_id] = {}
                
                for rule_name, rule_data in guild_rules.items():
                    pattern = rule_data.get("pattern", "")
                    if pattern:
                        try:
                            pattern_text, flags_letters = _parse_pattern_and_flags(pattern)
                            compiled = _compile_with_flags(pattern_text, flags_letters)
                            
                            regex_settings_by_guild[guild_id][rule_name] = {
                                "pattern": pattern,
                                "compiled": compiled,
                                "channels": set(rule_data.get("channels", [])),
                                "exempt_users": set(rule_data.get("exempt_users", [])),
                                "exempt_roles": set(rule_data.get("exempt_roles", []))
                            }
                        except Exception as e:
                            print(f"[SECURITY] Warning: Could not compile regex pattern '{pattern}': {e}")
            except ValueError:
                print(f"[SECURITY] Warning: Invalid guild ID in regex settings: {guild_id_str}")
        
        # Spam Settings
        spam_data = settings_data.get("spam_rules_by_guild", {})
        spam_rules_by_guild.clear()
        for guild_id_str, guild_rules in spam_data.items():
            try:
                guild_id = int(guild_id_str)
            except ValueError:
                print(f"[SECURITY] Warning: Invalid guild ID in spam settings: {guild_id_str}")
                continue
            spam_rules_by_guild[guild_id] = {}
            for rule_name, rule_data in guild_rules.items():
                try:
                    label = str(rule_data.get("label", rule_name))
                    min_length = int(rule_data.get("min_length", 0))
                    similarity_threshold = float(rule_data.get("similarity_threshold", 0.0))
                    time_window = int(rule_data.get("time_window", 0))
                    message_count = int(rule_data.get("message_count", 0))
                    dm_message = str(rule_data.get("dm_message", ""))
                    notify_channel_id = rule_data.get("notify_channel_id")
                    if notify_channel_id is not None:
                        notify_channel_id = int(notify_channel_id)
                    mod_action_value = rule_data.get("mod_action")
                    if isinstance(mod_action_value, str):
                        mod_action_value = mod_action_value.lower()
                        if mod_action_value not in {"warn", "delete", "warnanddelete"}:
                            mod_action_value = None
                    else:
                        mod_action_value = None
                    # Load regex_pattern if present (None for similarity mode)
                    regex_pattern_value = rule_data.get("regex_pattern")
                    if regex_pattern_value is not None:
                        regex_pattern_value = str(regex_pattern_value)
                        # Validate the regex pattern
                        try:
                            re.compile(regex_pattern_value, re.IGNORECASE)
                        except re.error:
                            print(f"[SECURITY] Warning: Invalid regex pattern in rule '{rule_name}' for guild {guild_id_str}")
                            regex_pattern_value = None

                    spam_rules_by_guild[guild_id][rule_name] = {
                        "label": label,
                        "min_length": max(0, min_length),
                        "similarity_threshold": max(0.0, min(similarity_threshold, 1.0)),
                        "time_window": max(0, time_window),
                        "message_count": max(0, message_count),
                        "dm_message": dm_message,
                        "notify_channel_id": notify_channel_id,
                        "channels": set(rule_data.get("channels", [])),
                        "excluded_channels": set(rule_data.get("excluded_channels", [])),
                        "targeted_roles": set(rule_data.get("targeted_roles", [])),
                        "exempted_roles": set(rule_data.get("exempted_roles", [])),
                        "nonreply_only": _coerce_bool(rule_data.get("nonreply_only", False)),
                        "mod_action": mod_action_value,
                        "regex_pattern": regex_pattern_value,
                    }
                except Exception as e:
                    print(f"[SECURITY] Warning: Could not load spam rule '{rule_name}' for guild {guild_id_str}: {e}")
        
        # Verify button usage
        usage_data = settings_data.get("verify_button_usage", {})
        verify_button_usage.clear()
        for user_id_str, count in usage_data.items():
            try:
                user_id = int(user_id_str)
                verify_button_usage[user_id] = count
            except ValueError:
                print(f"[SECURITY] Warning: Invalid user ID in verify button usage: {user_id_str}")
        
        # Loading statistics
        timestamp = settings_data.get("timestamp", 0)
        if timestamp:
            load_time = datetime.fromtimestamp(timestamp)
            print(f"[SECURITY] Settings loaded successfully (saved: {load_time.strftime('%Y-%m-%d %H:%M:%S')})")
        else:
            print(f"[SECURITY] Settings loaded successfully")
        
        # Summary of loaded settings
        print(f"[SECURITY] Loaded settings summary:")
        print(f"  - No-avatar filter: {'ON' if no_avatar_filter_enabled else 'OFF'}")
        print(f"  - Account age filter: {'ON' if account_age_filter_enabled else 'OFF'}")
        print(f"  - Authorized IDs: {len(security_authorized_ids)}")
        print(f"  - Whitelist users: {len(security_whitelist_users)}")
        print(f"  - Captcha role ID: {captcha_verify_role_id}")
        print(f"  - Panel texts for {len(captcha_panel_texts)} guilds")
        print(f"  - Regex rules for {len(regex_settings_by_guild)} guilds")
        if spam_rules_by_guild:
            total_spam_rules = sum(len(rules) for rules in spam_rules_by_guild.values())
            print(f"  - Spam rules: {total_spam_rules} rules in {len(spam_rules_by_guild)} guilds")
        print(f"  - Verify button usage for {len(verify_button_usage)} users")

        return True

    except Exception as e:
        print(f"[SECURITY] Error loading settings: {e}")
        return False

# Record function: Adds or updates a record in the Excel file.
# The discord_id is converted to a string to prevent scientific notation.
def record_play(discord_id, discord_username, in_game_username, event_name):
    file_name = f"{event_name}_play_records.xlsx"
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        
        # Check if user already exists in the sheet
        user_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == str(discord_id):
                user_row = row_idx
                break
        
        if user_row:
            # Update existing record
            sheet.cell(row=user_row, column=1, value=str(discord_id))
            sheet.cell(row=user_row, column=2, value=discord_username)
            sheet.cell(row=user_row, column=3, value=in_game_username)
        else:
            # Add new record
            sheet.append([str(discord_id), discord_username, in_game_username])
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Discord ID", "Discord Username", "In-Game Username"])
        # Save the discord_id as a string.
        sheet.append([str(discord_id), discord_username, in_game_username])
    
    workbook.save(file_name)

# ============== UTILITY FUNCTIONS ==============

def _chunk_message_lines(lines: list[str], limit: int = 1900):
    """Yield message chunks that respect Discord's character limit."""
    chunk: list[str] = []
    length = 0
    for raw_line in lines:
        line = str(raw_line) if raw_line is not None else ""
        if len(line) > limit:
            if chunk:
                yield "\n".join(chunk)
                chunk = []
                length = 0
            for start in range(0, len(line), limit):
                yield line[start : start + limit]
            continue
        if not chunk:
            chunk.append(line)
            length = len(line)
            continue
        addition = 1 + len(line)
        if length + addition > limit:
            yield "\n".join(chunk)
            chunk = [line]
            length = len(line)
        else:
            chunk.append(line)
            length += addition
    if chunk:
        yield "\n".join(chunk)


def _chunk_text_message(content: str, limit: int = 2000) -> List[str]:
    """Split long text into <=limit sized chunks, preferring newline breaks."""
    if content is None:
        return []
    text = str(content)
    if len(text) <= limit:
        return [text]

    chunks: List[str] = []
    remaining = text
    while remaining:
        if len(remaining) <= limit:
            chunks.append(remaining)
            break

        split_index = remaining.rfind("\n", 0, limit)
        if split_index == -1 or split_index < limit // 2:
            split_index = limit
            chunk = remaining[:split_index]
            remaining = remaining[split_index:]
        else:
            chunk = remaining[:split_index]
            remaining = remaining[split_index + 1 :]

        chunks.append(chunk.rstrip())
        remaining = remaining.lstrip()

    return [chunk or "" for chunk in chunks]


async def _send_long_message(
    sender: Callable[..., Awaitable["discord.Message"]],
    content: str,
    *,
    view: Optional["discord.ui.View"] = None,
    **kwargs,
) -> None:
    """Send content via sender, splitting into chunks and attaching view to first chunk."""

    chunks = _chunk_text_message(content)
    if not chunks:
        return

    first = True
    for chunk in chunks:
        send_kwargs = dict(kwargs)
        if first and view is not None:
            send_kwargs["view"] = view
        await sender(chunk, **send_kwargs)
        first = False


def _coerce_bool(value) -> bool:
    """Coerce various truthy representations into a real boolean."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"1", "true", "yes", "on", "enable", "enabled"}:
            return True
        if lowered in {"0", "false", "no", "off", "disable", "disabled"}:
            return False
        return bool(lowered)
    if isinstance(value, (int, float)):
        return value != 0
    return bool(value)


def _parse_role_ids(env_value: str | None) -> Set[int]:
    """Parse comma-separated role IDs from environment variable."""
    ids: Set[int] = set()
    if not env_value:
        return ids
    for token in env_value.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            ids.add(int(token))
        except ValueError:
            print(f"⚠️  WARNING: Invalid role ID entry ignored: {token!r}")
    return ids

# Function to calculate string similarity for fuzzy matching
def _parse_pattern_and_flags(raw_text):
    text = (raw_text or "").strip()
    flags_letters_parts = []

    # Support trailing --flags imsx style
    m = re.search(r"\s--flags\s+([A-Za-z]+)\s*$", text)
    if m:
        flags_letters_parts.append(m.group(1))
        text = text[: m.start()].strip()

    # Support /pattern/flags style
    if len(text) >= 2 and text[0] == "/" and "/" in text[1:]:
        last_slash = text.rfind("/")
        body = text[1:last_slash]
        trailing = text[last_slash + 1 :].strip()
        if trailing and re.fullmatch(r"[A-Za-z]+", trailing):
            flags_letters_parts.append(trailing)
            # Unescape \/ to /
            body = body.replace("\\/", "/")
            text = body

    # Merge letters while preserving order and uniqueness
    seen = set()
    letters = ""
    for part in flags_letters_parts:
        for ch in part:
            cl = ch.lower()
            if cl not in seen:
                seen.add(cl)
                letters += cl

    return text, letters


def _compile_with_flags(pattern_text, flags_letters):
    flag_map = {
        "i": re.IGNORECASE,
        "m": re.MULTILINE,
        "s": re.DOTALL,
        "x": re.VERBOSE,
        "a": re.ASCII,
        "u": re.UNICODE,
        "l": re.LOCALE,
    }
    flags_value = 0
    for ch in flags_letters or "":
        flags_value |= flag_map.get(ch.lower(), 0)
    return _REGEX_ENGINE.compile(pattern_text, flags_value)

def string_similarity(s1, s2):
    # Convert both strings to lowercase for case-insensitive comparison
    s1 = s1.lower()
    s2 = s2.lower()
    
    # Calculate similarity ratio using difflib
    similarity = difflib.SequenceMatcher(None, s1, s2).ratio()
    return similarity

# Intent settings
intents = discord.Intents.default()
intents.members = True  # Required for member join events
try:
    intents.messages = True  # ensure message create events
except Exception:
    pass
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

# Global check: disable all commands in DMs (guild-only)
@bot.check
async def _block_dm_commands(ctx: commands.Context) -> bool:
    # Return False for DMs so commands are ignored
    return ctx.guild is not None

# Global variables
events = {}
usage_counts = {}
event_nickname_counts = {}
event_nickname_limit = {}  # For same nickname limit

# Captcha verification settings
captcha_verify_role_id = None  # Role to grant upon successful captcha

# Customizable verification panel text per guild
captcha_panel_texts = {}  # guild_id -> {"title": str, "description": str, "image": str}

# Default panel text
DEFAULT_PANEL_TITLE = "Verification Panel"
DEFAULT_PANEL_DESCRIPTION = (
    "Server access requires verification.\n"
    "Click 'Verify' to receive a visual CAPTCHA challenge.\n"
    "Enter the displayed code using the 'Enter Code' button to complete verification."
)

# Rate limiting for captcha requests
captcha_rate_limits = defaultdict(list)  # user_id -> [timestamp, timestamp, ...]
CAPTCHA_RATE_LIMIT = 3  # Max 3 requests per minute
CAPTCHA_RATE_WINDOW = 60  # 60 seconds window

# Active captcha sessions to prevent spam
active_captcha_sessions = set()  # Set of user_ids currently processing captcha

# Verify button usage tracking (user_id -> interaction_count)
verify_button_usage = defaultdict(int)  # Tracks how many times each user clicked verify
VERIFY_MAX_ATTEMPTS = 10  # Maximum verify attempts per user

# Rate limit message tracking for CAPTCHA rate limits (user_id -> last_message_time)
captcha_rate_limit_messages = defaultdict(float)  # Tracks when captcha rate limit message was last sent
CAPTCHA_RATE_LIMIT_MESSAGE_COOLDOWN = 60  # Show captcha rate limit message once per minute

# Regex moderation settings per guild
# Structure: { guild_id: { name: {"pattern": str, "compiled": Pattern, "channels": set[int], "exempt_users": set[int], "exempt_roles": set[int]} } }
regex_settings_by_guild = {}

# Spam moderation settings per guild
# Structure: { guild_id: { name: {"min_length": int, "similarity_threshold": float, "time_window": int, "message_count": int, "dm_message": str, "notify_channel_id": int, "channels": set[int], "nonreply_only": bool, "mod_action": str | None} } }
spam_rules_by_guild = {}

# Security Authorization
# Load security role ID from environment variable for better security
security_authorized_role_ids: Set[int] = _parse_role_ids(os.getenv("SECURITY_MANAGER_ROLE_ID"))
if not security_authorized_role_ids:
    print("⚠️  WARNING: SECURITY_MANAGER_ROLE_ID environment variable not set or invalid!")
    print("Security commands will only work with manually added IDs via !securityauthorizedadd")
    print("To set default security roles: export SECURITY_MANAGER_ROLE_ID='id1,id2,...'")

security_authorized_ids = set()

# Whitelist for security filters (users on whitelist bypass noavatar and account age filters)
security_whitelist_users: Set[int] = set()

# Security limits and audit
MAX_SECURITY_AUTHORIZED_USERS = 4  # Maximum number of authorized users/roles
security_audit_log = []  # Store security actions for audit

# Rate limiting for critical operations
command_rate_limits = defaultdict(list)  # user_id -> [timestamp, timestamp, ...]
SECURITY_COMMAND_RATE_LIMIT = 5  # Max 5 security commands per minute
SECURITY_COMMAND_RATE_WINDOW = 60  # 60 seconds window

# Rate limit message tracking for security commands
security_rate_limit_messages = defaultdict(float)
SECURITY_RATE_LIMIT_MESSAGE_COOLDOWN = 30  # Show rate limit message once per 30 seconds

def is_security_authorized(ctx):
    if security_authorized_role_ids and any(role.id in security_authorized_role_ids for role in ctx.author.roles):
        return True
    if ctx.author.id in security_authorized_ids:
        return True
    for role in ctx.author.roles:
        if role.id in security_authorized_ids:
            return True
    return False

def _check_security_command_rate_limit(user_id: int) -> bool:
    """Check if user is within rate limit for security commands"""
    current_time = time.time()
    user_requests = command_rate_limits[user_id]
    
    # Remove old requests outside the window
    user_requests[:] = [req_time for req_time in user_requests if current_time - req_time < SECURITY_COMMAND_RATE_WINDOW]
    
    # Check if user has exceeded rate limit
    if len(user_requests) >= SECURITY_COMMAND_RATE_LIMIT:
        return False
    
    return True

def _add_security_command_rate_limit_request(user_id: int):
    """Add a request to the security command rate limit tracker"""
    current_time = time.time()
    command_rate_limits[user_id].append(current_time)

async def _handle_security_rate_limit(ctx, command_name: str) -> bool:
    """Handle rate limiting for security commands. Returns True if rate limited."""
    user_id = ctx.author.id
    
    if not _check_security_command_rate_limit(user_id):
        current_time = time.time()
        last_rate_limit_message = security_rate_limit_messages[user_id]
        
        # Show rate limit message only once per cooldown period
        if current_time - last_rate_limit_message >= SECURITY_RATE_LIMIT_MESSAGE_COOLDOWN:
            security_rate_limit_messages[user_id] = current_time
            await ctx.send(f"⏰ **Rate limit exceeded!** You can only use {SECURITY_COMMAND_RATE_LIMIT} security commands per minute. Please wait and try again.")
        
        # Delete the command message to reduce spam
        try:
            await ctx.message.delete()
        except discord.Forbidden:
            print(f"[SECURITY] Bot lacks permission to delete command message")
        except discord.NotFound:
            print(f"[SECURITY] Command message already deleted")
        except discord.HTTPException as e:
            print(f"[SECURITY] HTTP error deleting command message: {e}")
        except Exception as e:
            print(f"[SECURITY] Unexpected error deleting command message: {e}")
        
        return True  # Rate limited
    
    # Add to rate limit tracker
    _add_security_command_rate_limit_request(user_id)
    return False  # Not rate limited

# Play Event Authorization
# Load play role ID from environment variable
play_authorized_role_id = int(os.getenv("PLAY_MANAGER_ROLE_ID", "0")) or None
play_authorized_ids = set()

def is_play_authorized(ctx):
    if play_authorized_role_id in [role.id for role in ctx.author.roles]:
        return True
    if ctx.author.id in play_authorized_ids:
        return True
    for role in ctx.author.roles:
        if role.id in play_authorized_ids:
            return True
    return False

# Role IDs allowed to use the button
allowed_role_ids = [1346571303657930804]   # The role ID defined for button usage

# Global Security Filter Variables
no_avatar_filter_enabled = False
no_avatar_action = None
no_avatar_timeout_duration = None

account_age_filter_enabled = False
account_age_min_days = None
account_age_action = None
account_age_timeout_duration = None

# Spam violation statistics configuration
SPAM_STATS_FILE = Path(__file__).with_name("spam_violation_stats.json")
SPAM_AGGREGATE_WINDOWS = [
    ("24h", 1),
    ("7d", 7),
    ("30d", 30),
    ("90d", 90),
    ("120d", 120),
    ("180d", 180),
    ("360d", 360),
]
MAX_SPAM_AGGREGATE_DAYS = 360
spam_violation_stats = {}
spam_stats_loaded = False
spam_stats_lock = asyncio.Lock()

# Commonly used time windows supported out-of-the-box (value in seconds)
SPAM_RULE_PREDEFINED_WINDOWS = {
    "24h": 24 * 3600,
    "7d": 7 * 86400,
    "30d": 30 * 86400,
    "60d": 60 * 86400,
    "90d": 90 * 86400,
    "120d": 120 * 86400,
    "180d": 180 * 86400,
    "360d": 360 * 86400,
}

# Runtime spam tracking (not persisted)
# Key: (guild_id, user_id) -> list[{"timestamp": float, "content": str}]
spam_message_history = defaultdict(list)

# Last trigger timestamps to prevent duplicate alerts within the window
# Key: (guild_id, user_id, rule_name) -> float
spam_rule_trigger_log = {}

# ============== SPAM & REGEX HELPER FUNCTIONS ==============

_WORD_TOKEN_PATTERN = re.compile(r"\w+")

def _extract_word_tokens(text: str) -> list[str]:
    """Return lowercase word tokens extracted from text."""
    if not text:
        return []
    return [token for token in _WORD_TOKEN_PATTERN.findall(text.lower()) if token]

def _token_multiset_similarity(tokens1: list[str], tokens2: list[str]) -> float:
    """Compute multiset Jaccard similarity between two token lists."""
    if not tokens1 or not tokens2:
        return 0.0

    counts1 = Counter(tokens1)
    counts2 = Counter(tokens2)
    intersection = sum(min(counts1[token], counts2[token]) for token in counts1.keys())
    union = sum(max(counts1.get(token, 0), counts2.get(token, 0)) for token in set(counts1) | set(counts2))
    if union == 0:
        return 0.0
    return intersection / union

def _is_message_reply(message: discord.Message) -> bool:
    """Return True if message is a Discord reply (covers uncached targets)."""
    ref = getattr(message, "reference", None)
    if ref is not None:
        if getattr(ref, "resolved", None) is not None:
            return True
        if any(getattr(ref, attr, None) for attr in ("message_id", "channel_id", "guild_id")):
            return True
        return True
    try:
        if message.type == discord.MessageType.reply:
            return True
    except AttributeError:
        pass
    return False

# Security: Safe regex search with timeout to prevent ReDoS attacks
def _safe_regex_search(compiled_pattern, text, timeout_seconds=1):
    """Safely search regex with timeout to prevent ReDoS (Regular Expression Denial of Service) attacks"""
    if not text:
        return None
    
    # Limit text length to prevent memory issues
    MAX_TEXT_LENGTH = 10000
    if len(text) > MAX_TEXT_LENGTH:
        text = text[:MAX_TEXT_LENGTH]
    
    result = [None]
    exception = [None]
    
    def search_worker():
        try:
            result[0] = compiled_pattern.search(text)
        except Exception as e:
            exception[0] = e
    
    # Use threading for timeout (signal doesn't work well with Discord.py)
    thread = threading.Thread(target=search_worker, daemon=True)
    thread.start()
    thread.join(timeout=timeout_seconds)
    
    if thread.is_alive():
        # Timeout occurred - potential ReDoS attack
        print(f"[SECURITY] Regex timeout detected - potential ReDoS attack blocked")
        return None
    
    if exception[0]:
        print(f"[SECURITY] Regex error: {exception[0]}")
        return None
        
    return result[0]

def _collect_regex_text_blocks(
    message: discord.Message,
    *,
    _seen: Optional[Set[int]] = None
) -> List[str]:
    """Return textual fragments that should be scanned by regex rules."""
    blocks: List[str] = []
    seen: Set[int] = _seen or set()

    message_id = getattr(message, "id", None)
    if message_id is not None:
        if message_id in seen:
            return blocks
        seen.add(message_id)

    content = getattr(message, "content", None)
    if content:
        blocks.append(content)

    # Also include system-rendered text (forwarded/announcement/system messages)
    system_content = getattr(message, "system_content", None)
    if isinstance(system_content, str) and system_content.strip():
        if system_content != content:
            blocks.append(system_content)

    # Include attachment URLs
    try:
        for att in getattr(message, "attachments", []) or []:
            url = getattr(att, "url", None)
            if url:
                blocks.append(str(url))
            proxy_url = getattr(att, "proxy_url", None)
            if proxy_url and proxy_url != url:
                blocks.append(str(proxy_url))
    except Exception:
        pass

    for embed in getattr(message, "embeds", []) or []:
        title = getattr(embed, "title", None)
        if title:
            blocks.append(title)
        description = getattr(embed, "description", None)
        if description:
            blocks.append(description)
        try:
            e_url = getattr(embed, "url", None)
            if e_url:
                blocks.append(str(e_url))
            # Skip CDN URLs (thumbnail, image, video) - these are auto-generated by Discord
            # and cause false positives with whitelist domains like x.com
        except Exception:
            pass
        for field in getattr(embed, "fields", []) or []:
            field_name = getattr(field, "name", None)
            field_value = getattr(field, "value", None)
            if field_name:
                blocks.append(field_name)
            if field_value:
                blocks.append(field_value)
        # Skip footer - Discord metadata
        author = getattr(embed, "author", None)
        if author:
            author_name = getattr(author, "name", None)
            if author_name:
                blocks.append(author_name)
            author_url = getattr(author, "url", None)
            if author_url:
                blocks.append(str(author_url))

    reference = getattr(message, "reference", None)
    if reference:
        resolved = getattr(reference, "resolved", None)
        cached = getattr(reference, "cached_message", None)
        target = None
        if isinstance(resolved, discord.Message):
            target = resolved
        elif isinstance(cached, discord.Message):
            target = cached
        if target is not None:
            blocks.extend(_collect_regex_text_blocks(target, _seen=seen))

    # Forwarded message snapshots (Discord API v2.5+)
    try:
        snapshots = getattr(message, "message_snapshots", None)
        if snapshots:
            for snap in snapshots or []:
                s_content = getattr(snap, "content", None)
                if s_content:
                    blocks.append(s_content)
                for emb in getattr(snap, "embeds", []) or []:
                    t = getattr(emb, "title", None)
                    d = getattr(emb, "description", None)
                    if t:
                        blocks.append(t)
                    if d:
                        blocks.append(d)
                    try:
                        eurl = getattr(emb, "url", None)
                        if eurl:
                            blocks.append(str(eurl))
                        # Skip CDN URLs (thumbnail, image, video) for snapshots too
                        # Include author URL for embedded tweets/posts
                        auth = getattr(emb, "author", None)
                        if auth is not None:
                            auth_url = getattr(auth, "url", None)
                            if auth_url:
                                blocks.append(str(auth_url))
                    except Exception:
                        pass
                for att in getattr(snap, "attachments", []) or []:
                    au = getattr(att, "url", None)
                    if au:
                        blocks.append(str(au))
    except Exception:
        pass

    return [block for block in blocks if isinstance(block, str) and block.strip()]

# Helper function for regex moderation (shared by on_message and on_message_edit)
async def _check_message_against_regex(message: discord.Message):
    """Check message against regex rules and delete if it matches"""
    if message.guild is None:
        return

    # Apply target channel rules regardless of the source/origin.
    # Only skip our own bot's messages to prevent loops; check everything else (including webhooks/other bots).
    try:
        if bot.user and message.author and message.author.id == bot.user.id:
            return
    except Exception:
        pass

    text_blocks = _collect_regex_text_blocks(message)
    if DEBUG_MODE:
        try:
            snaps = getattr(message, 'message_snapshots', None)
            print(
                f"[REGEX_SCAN] blocks={len(text_blocks)} content_len={len(message.content or '')} "
                f"embeds={len(getattr(message,'embeds',[]) or [])} atts={len(getattr(message,'attachments',[]) or [])} "
                f"snapshots={len(snaps) if snaps is not None else 0}",
                flush=True,
            )
        except Exception:
            pass
    if not text_blocks:
        return

    # Debug logging for message scanning
    if DEBUG_MODE:
        try:
            print(f"[DEBUG] Scanning message in channel {message.channel.id} ({getattr(message.channel, 'name', '?')}) | author_bot={getattr(message.author, 'bot', None)} webhook_id={getattr(message, 'webhook_id', None)} flags={getattr(getattr(message,'flags',None),'value', None)}")
        except Exception:
            pass

    guild_rules = regex_settings_by_guild.get(message.guild.id)
    if not guild_rules:
        return

    channel_id = message.channel.id
    # For threads (forum posts), also check parent channel ID
    parent_id = None
    if isinstance(message.channel, discord.Thread):
        parent_id = message.channel.parent_id
    
    for rule in guild_rules.values():
        channels = rule.get("channels", set())
        compiled = rule.get("compiled")
        if not compiled or not channels:
            continue
        # Check if message is in a monitored channel or thread within a monitored channel
        if channel_id not in channels:
            if parent_id is None or parent_id not in channels:
                if DEBUG_MODE:
                    try:
                        print(f"[DEBUG] Message not in monitored channels. Channel: {channel_id}, Parent: {parent_id}, Monitored count: {len(channels)}", flush=True)
                    except Exception:
                        pass
                continue

        if not any(_safe_regex_search(compiled, text) for text in text_blocks):
            continue

        exempt_users = rule.get("exempt_users", set())
        if message.author.id in exempt_users:
            continue
        exempt_roles = rule.get("exempt_roles", set())
        author_roles = getattr(message.author, "roles", [])
        if any(r.id in exempt_roles for r in author_roles):
            continue

        try:
            await message.delete()
        except discord.Forbidden:
            print(f"[SECURITY] Bot lacks permission to delete message in {message.channel}")
        except discord.NotFound:
            print(f"[SECURITY] Message already deleted in {message.channel}")
        except discord.HTTPException as e:
            print(f"[SECURITY] HTTP error deleting message: {e}")
        except Exception as e:
            print(f"[SECURITY] Unexpected error deleting message: {e}")
        break

async def _check_message_against_spam_rules(message: discord.Message):
    """Check message against custom spam rules and apply configured actions"""
    if message.author.bot:
        return
    if message.guild is None:
        return

    # Skip security managers / authorized users
    author_roles = getattr(message.author, "roles", []) or []
    if security_authorized_role_ids and any(r.id in security_authorized_role_ids for r in author_roles):
        return
    if message.author.id in security_authorized_ids:
        return
    if any(r.id in security_authorized_ids for r in author_roles):
        return

    guild_rules = spam_rules_by_guild.get(message.guild.id)
    if not guild_rules:
        return

    content = message.content or ""
    if not content:
        return

    content_tokens = _extract_word_tokens(content)

    now = time.time()
    history_key = (message.guild.id, message.author.id)
    user_history = spam_message_history[history_key]

    is_reply = _is_message_reply(message)

    max_window = 0
    for rule in guild_rules.values():
        window = rule.get("time_window", 0)
        if window > max_window:
            max_window = window

    if max_window > 0:
        user_history[:] = [entry for entry in user_history if now - entry["timestamp"] <= max_window]
    else:
        user_history.clear()

    user_history.append({
        "timestamp": now,
        "content": content,
        "is_reply": is_reply,
        "channel_id": message.channel.id,
        "tokens": content_tokens,
    })

    for name_key, rule in guild_rules.items():
        # Check excluded channels first
        excluded_channels = rule.get("excluded_channels", set())
        if excluded_channels and message.channel.id in excluded_channels:
            continue

        # Check included channels
        channels = rule.get("channels", set())
        if channels and message.channel.id not in channels:
            continue

        # Check exempted roles first
        exempted_roles = rule.get("exempted_roles", set())
        if exempted_roles:
            user_role_ids = {role.id for role in message.author.roles}
            if user_role_ids & exempted_roles:  # User has at least one exempted role
                continue

        # Check targeted roles
        targeted_roles = rule.get("targeted_roles", set())
        if targeted_roles:
            user_role_ids = {role.id for role in message.author.roles}
            if not (user_role_ids & targeted_roles):  # User has none of the targeted roles
                continue

        nonreply_only = rule.get("nonreply_only", False)
        # Skip this rule for reply messages if nonreply_only is enabled
        if nonreply_only and is_reply:
            continue

        min_length = rule.get("min_length", 0)
        if min_length and len(content) <= min_length:
            continue

        time_window = rule.get("time_window", 0)
        if time_window <= 0:
            continue

        message_count = rule.get("message_count", 0)
        if message_count <= 1:
            continue

        # Check if this is a regex-based rule or similarity-based rule
        regex_pattern_str = rule.get("regex_pattern")
        similarity_threshold = rule.get("similarity_threshold", 0.0)

        # For similarity mode, require similarity_threshold > 0
        # For regex mode, regex_pattern must be set
        if not regex_pattern_str and similarity_threshold <= 0:
            continue

        relevant_messages = [
            entry
            for entry in user_history
            if now - entry["timestamp"] <= time_window
            and (
                not channels
                or entry.get("channel_id") is None
                or entry["channel_id"] in channels
            )
        ]
        # Filter out reply messages for nonreply_only rules
        if nonreply_only:
            relevant_messages = [
                entry for entry in relevant_messages
                if not _coerce_bool(entry.get("is_reply", False))
            ]
        if len(relevant_messages) < message_count:
            continue

        if regex_pattern_str:
            # REGEX MODE: Count messages that match the regex pattern
            try:
                compiled_pattern = re.compile(regex_pattern_str, re.IGNORECASE)
            except re.error:
                continue  # Invalid regex, skip this rule

            matching_count = 0
            for entry in relevant_messages:
                entry_content = entry.get("content", "")
                if compiled_pattern.search(entry_content):
                    matching_count += 1
            if matching_count >= message_count:
                await _handle_spam_rule_trigger(message, name_key, rule)
                continue
        else:
            # SIMILARITY MODE: Original behavior
            similar_count = 0
            for entry in relevant_messages:
                entry_content = entry.get("content", "")
                char_ratio = SequenceMatcher(None, content, entry_content).ratio() if entry_content else 0.0

                entry_tokens = entry.get("tokens")
                if entry_tokens is None:
                    entry_tokens = _extract_word_tokens(entry_content)
                    entry["tokens"] = entry_tokens

                token_ratio = _token_multiset_similarity(content_tokens, entry_tokens)
                ratio = max(char_ratio, token_ratio)
                if ratio >= similarity_threshold:
                    similar_count += 1
            if similar_count >= message_count:
                await _handle_spam_rule_trigger(message, name_key, rule)
                continue

async def _handle_spam_rule_trigger(message: discord.Message, rule_key: str, rule: dict):
    """Execute actions when a spam rule is triggered"""
    guild_id = message.guild.id
    user_id = message.author.id
    now = time.time()

    cooldown = max(rule.get("trigger_cooldown", 0), 0)
    last_trigger = spam_rule_trigger_log.get((guild_id, user_id, rule_key))
    if last_trigger and now - last_trigger < cooldown:
        return

    spam_rule_trigger_log[(guild_id, user_id, rule_key)] = now

    await record_spam_violation(guild_id, user_id, rule_key, label=rule.get("label", rule_key))

    mod_action = (rule.get("mod_action") or "").lower()
    dm_message = rule.get("dm_message")
    should_dm = bool(dm_message) and mod_action in {"", "warn", "warnanddelete"}
    should_delete = mod_action in {"delete", "warnanddelete"}

    message_content_snapshot = message.content or ""

    dm_sent = False
    dm_error: str | None = None
    if should_dm:
        try:
            await message.author.send(dm_message)
            dm_sent = True
        except discord.Forbidden:
            dm_error = "Forbidden"
            print(f"[SECURITY] Unable to DM user {user_id} for spam rule '{rule_key}' (forbidden)")
        except discord.HTTPException as e:
            dm_error = f"HTTP {getattr(e, 'status', 'error')}"
            print(f"[SECURITY] HTTP error sending DM: {e}")
        except Exception as e:
            dm_error = "Unexpected error"
            print(f"[SECURITY] Unexpected error sending DM: {e}")

    delete_success = False
    delete_error: str | None = None
    if should_delete:
        try:
            await message.delete()
            delete_success = True
        except discord.NotFound:
            delete_error = "Message already deleted"
            print(f"[SECURITY] Message already deleted when applying spam rule '{rule_key}'")
        except discord.Forbidden:
            delete_error = "Missing permissions"
            print(f"[SECURITY] Missing permissions to delete message for spam rule '{rule_key}'")
        except discord.HTTPException as e:
            delete_error = f"HTTP {getattr(e, 'status', 'error')}"
            print(f"[SECURITY] HTTP error deleting message for spam rule '{rule_key}': {e}")
        except Exception as e:
            delete_error = "Unexpected error"
            print(f"[SECURITY] Unexpected error deleting message for spam rule '{rule_key}': {e}")

    notify_channel_id = rule.get("notify_channel_id")
    if notify_channel_id:
        channel = message.guild.get_channel(notify_channel_id)
        if channel:
            label = rule.get("label", rule_key)
            window_seconds = rule.get("time_window", 0)
            try:
                preview = message_content_snapshot[:1500].strip()
                if not preview:
                    preview = "(no content)"

                if mod_action == "warn":
                    action_summary = "Warn via DM"
                elif mod_action == "delete":
                    action_summary = "Delete message"
                elif mod_action == "warnanddelete":
                    action_summary = "Warn via DM & delete message"
                else:
                    action_summary = "Warn via DM"

                outcome_bits = []
                if should_dm:
                    outcome_bits.append("DM sent" if dm_sent else f"DM failed{f' ({dm_error})' if dm_error else ''}")
                if should_delete:
                    outcome_bits.append("Message deleted" if delete_success else f"Delete failed{f' ({delete_error})' if delete_error else ''}")
                outcome_text = ", ".join(outcome_bits) if outcome_bits else "N/A"

                await channel.send(
                    f"⚠️ Spam rule `{label}` triggered by {message.author.mention} in {message.channel.mention}.\n"
                    f"Window: {window_seconds} seconds | Similarity ≥ {int(rule.get('similarity_threshold', 0.0) * 100)}% | Count ≥ {rule.get('message_count', 0)}\n"
                    f"Action: {action_summary} | Outcome: {outcome_text}\n"
                    f"Recent message:\n```{preview}```"
                )
            except discord.HTTPException as e:
                print(f"[SECURITY] HTTP error notifying channel {notify_channel_id}: {e}")
            except Exception as e:
                print(f"[SECURITY] Unexpected error notifying channel {notify_channel_id}: {e}")
        else:
            print(f"[SECURITY] Notification channel {notify_channel_id} not found for spam rule '{rule_key}'")

# Message moderation via regex
@bot.event
async def on_message(message: discord.Message):
    # Debug logs (only when DEBUG_MODE is enabled)
    if DEBUG_MODE:
        try:
            flags = getattr(message, "flags", None)
            is_cross = bool(getattr(flags, "is_crossposted", False) or getattr(flags, "crossposted", False)) if flags else False
            is_webhook = getattr(message, "webhook_id", None) is not None
            is_bot_author = bool(getattr(getattr(message, "author", None), "bot", False))
            own_bot = False
            try:
                own_bot = bool(bot.user and message.author and message.author.id == bot.user.id)
            except Exception:
                own_bot = False
            if is_cross or is_webhook:
                ch_name = getattr(message.channel, "name", "?")
                print(f"[FORWARD_LOG] guild={getattr(getattr(message,'guild',None),'id',None)} ch={message.channel.id}/{ch_name} webhook={is_webhook} cross={is_cross} len={len(message.content or '')} sys={bool(getattr(message,'system_content', None))}", flush=True)
            if (is_bot_author and not own_bot) or is_webhook:
                ch_name = getattr(message.channel, "name", "?")
                print(
                    f"[BOT_OR_WEBHOOK_LOG] guild={getattr(getattr(message,'guild',None),'id',None)} ch={message.channel.id}/{ch_name} "
                    f"author_bot={is_bot_author} own_bot={own_bot} webhook={is_webhook} type={getattr(message,'type', None)}",
                    flush=True)
        except Exception:
            pass

    # Let command processor run only in guilds
    if message.guild and isinstance(bot.command_prefix, str) and message.content.startswith(bot.command_prefix):
        await bot.process_commands(message)
        return
    
    # Check message against regex rules
    await _check_message_against_regex(message)

    # Check custom spam rules
    await _check_message_against_spam_rules(message)

# Message edit moderation via regex
@bot.event
async def on_message_edit(before: discord.Message, after: discord.Message):
    """Check edited messages against regex rules"""
    # Only check the edited message (after)
    await _check_message_against_regex(after)

# Button interaction handler - Add this to fix the interaction failed issue
@bot.event
async def on_interaction(interaction):
    if interaction.type == discord.InteractionType.component:
        custom_id = interaction.data.get("custom_id", "")
        
        # Handle CAPTCHA verification button
        if custom_id == "captcha_verify_button":
            user_id = interaction.user.id
            
            # Check if user has exceeded verify attempts - COMPLETELY SILENT
            if verify_button_usage[user_id] >= VERIFY_MAX_ATTEMPTS:
                # Silently ignore - no response at all
                return
            
            # Check if verification role is set
            if captcha_verify_role_id is None:
                await interaction.response.send_message(
                    "Verification role is not set. Please contact an administrator.",
                    ephemeral=True,
                )
                return

            # Check rate limiting BEFORE incrementing usage count
            if not _check_captcha_rate_limit(user_id):
                current_time = time.time()
                last_rate_limit_message = captcha_rate_limit_messages[user_id]
                
                # Show rate limit message only once per minute
                if current_time - last_rate_limit_message >= CAPTCHA_RATE_LIMIT_MESSAGE_COOLDOWN:
                    captcha_rate_limit_messages[user_id] = current_time
                    await interaction.response.send_message(
                        f"⏰ Rate limit exceeded. You can only request {CAPTCHA_RATE_LIMIT} captchas per minute. Please wait and try again.",
                        ephemeral=True,
                    )
                # If rate limit message was sent recently, silently ignore
                # IMPORTANT: Don't increment usage count when rate limited
                return
            
            # Only increment usage count AFTER passing rate limit check
            verify_button_usage[user_id] += 1

            # Remove any existing active session to allow fresh captcha
            if user_id in active_captcha_sessions:
                active_captcha_sessions.discard(user_id)
                print(f"[captcha] Removed existing session for user {user_id} to generate fresh captcha")

            # Already has role?
            member = interaction.user
            if isinstance(member, discord.Member):
                if any(r.id == captcha_verify_role_id for r in getattr(member, "roles", [])):
                    await interaction.response.send_message(
                        "You are already verified.", ephemeral=True
                    )
                    return

            # Add to active sessions
            active_captcha_sessions.add(user_id)

            try:
                # Generate fresh captcha code each time
                code = _generate_captcha_code()
                print(f"[captcha] Generated FRESH code: {code} for user {user_id} (attempt #{verify_button_usage[user_id]}), PIL available: {_PIL_AVAILABLE}")

                # Add to rate limit tracker only when captcha is successfully generated
                _add_captcha_rate_limit_request(user_id)

                # Defer first, then send image
                await interaction.response.defer(ephemeral=True)
                
                # Text image CAPTCHA
                if _PIL_AVAILABLE:
                    try:
                        img_bytes = _create_text_image(code)
                        file = discord.File(io.BytesIO(img_bytes), filename="captcha.png")
                        message_text = f"🔐 **Security Verification**\n\nPlease read the code from the image below and click 'Enter Code' to input it.\n\n**Attempt: {verify_button_usage[user_id]}/{VERIFY_MAX_ATTEMPTS}**"
                        view = CaptchaCodeEntryView(expected_code=code, verify_role_id=captcha_verify_role_id, user_id=user_id)
                        await interaction.followup.send(content=message_text, file=file, view=view, ephemeral=True)
                        return
                    except Exception as e:
                        print(f"[captcha] Text image creation failed: {e}")

                # Fallback: Simple text
                message_text = f"🔐 **Security Verification**\n\n**Code: `{code}`**\n\nPlease enter the code above by clicking 'Enter Code'.\n\n**Attempt: {verify_button_usage[user_id]}/{VERIFY_MAX_ATTEMPTS}**"
                view = CaptchaCodeEntryView(expected_code=code, verify_role_id=captcha_verify_role_id, user_id=user_id)
                await interaction.followup.send(content=message_text, view=view, ephemeral=True)

            except Exception as e:
                print(f"[captcha] Error in verify_button: {e}")
                # Remove from active sessions on error
                active_captcha_sessions.discard(user_id)
                try:
                    await interaction.followup.send("An error occurred while generating captcha. Please try again.", ephemeral=True)
                except Exception:
                    pass
            return
            
        # Handle CAPTCHA code entry button  
        if custom_id == "captcha_enter_code":
            # This should be handled by the CaptchaCodeEntryView
            return
            
        if custom_id.startswith("play_button_"):
            event_name = custom_id.replace("play_button_", "")
            
            # Check if event exists
            if event_name not in events:
                await interaction.response.send_message("This event no longer exists.", ephemeral=True)
                return
            
            # Check if user has allowed role
            has_allowed_role = False
            if not allowed_role_ids:  # If no roles are specified, allow all
                has_allowed_role = True
            else:
                for role in interaction.user.roles:
                    if role.id in allowed_role_ids:
                        has_allowed_role = True
                        break
            
            if not has_allowed_role:
                await interaction.response.send_message("You don't have the required role to use this button.", ephemeral=True)
                return
            
            # Check user limit if set
            user_limit = None
            for role in interaction.user.roles:
                if role.id in events[event_name]["limits"]:
                    user_limit = events[event_name]["limits"][role.id]
                    break
            
            if user_limit is not None:
                usage_key = f"{event_name}_{interaction.user.id}"
                if usage_counts.get(usage_key, 0) >= user_limit:
                    await interaction.response.send_message(
                        f"You've reached your limit of {user_limit} interactions for this event.",
                        ephemeral=True
                    )
                    return
            
            # Create and send the modal
            class NicknameModal(discord.ui.Modal):
                def __init__(self, event_name, nickname_limit):
                    super().__init__(title=f"Register for {event_name}")
                    self.event_name = event_name
                    self.nickname_limit = nickname_limit
                    
                    self.nickname = discord.ui.TextInput(
                        label="Your In-Game Username",
                        placeholder="Enter your in-game username here...",
                        min_length=3,
                        max_length=32
                    )
                    self.add_item(self.nickname)
                
                async def on_submit(self, interaction):
                    in_game_username = self.nickname.value.strip()
                    
                    # Check same nickname limit if enabled
                    if self.nickname_limit is not None:
                        if in_game_username in event_nickname_counts.get(self.event_name, {}):
                            if event_nickname_counts[self.event_name][in_game_username] >= self.nickname_limit:
                                await interaction.response.send_message(
                                    f"This in-game username has already been registered {self.nickname_limit} times.",
                                    ephemeral=True
                                )
                                return
                        
                        # Update nickname count
                        if self.event_name not in event_nickname_counts:
                            event_nickname_counts[self.event_name] = {}
                        event_nickname_counts[self.event_name][in_game_username] = event_nickname_counts[self.event_name].get(in_game_username, 0) + 1
                    
                    # Record the play
                    record_play(interaction.user.id, str(interaction.user), in_game_username, self.event_name)
                    
                    # Update usage count if limits are set
                    if events[self.event_name]["limits"]:
                        usage_key = f"{self.event_name}_{interaction.user.id}"
                        usage_counts[usage_key] = usage_counts.get(usage_key, 0) + 1
                    
                    # Send confirmation message
                    await interaction.response.send_message(
                        f"Successfully registered for {self.event_name} with username: {in_game_username}",
                        ephemeral=True
                    )
                    
                    # Send link if available
                    event_link = events[self.event_name].get("link")
                    event_password = events[self.event_name].get("password")
                    
                    if event_link:
                        try:
                            link_message = f"Link for {self.event_name}: {event_link}"
                            if event_password:
                                link_message += f"\nPassword: {event_password}"
                            
                            await interaction.followup.send(link_message, ephemeral=True)
                        except Exception as e:
                            print(f"Error sending link: {e}")
            
            # Send the modal
            if event_name in event_nickname_limit:
                modal = NicknameModal(event_name, event_nickname_limit[event_name])
            else:
                modal = NicknameModal(event_name, None)
            
            await interaction.response.send_modal(modal)

# on_member_join event (Security Filters)
@bot.event
async def on_member_join(member):
    # Check if user is on whitelist (bypass all security filters)
    if member.id in security_whitelist_users:
        return
    
    if no_avatar_filter_enabled:
        if member.avatar is None:
            try:
                if no_avatar_action == "ban":
                    await member.ban(reason="No avatar provided")
                elif no_avatar_action == "kick":
                    await member.kick(reason="No avatar provided")
                elif no_avatar_action == "timeout":
                    timeout_duration = timedelta(minutes=no_avatar_timeout_duration)
                    until = discord.utils.utcnow() + timeout_duration
                    await member.edit(timeout=until, reason="No avatar provided")
            except Exception as e:
                print("No-avatar filter error:", e)
    if account_age_filter_enabled:
        account_age = (discord.utils.utcnow() - member.created_at).days
        if account_age < account_age_min_days:
            try:
                if account_age_action == "ban":
                    await member.ban(reason="Account age insufficient")
                elif account_age_action == "kick":
                    await member.kick(reason="Account age insufficient")
                elif account_age_action == "timeout":
                    timeout_duration = timedelta(minutes=account_age_timeout_duration)
                    until = discord.utils.utcnow() + timeout_duration
                    await member.edit(timeout=until, reason="Account age insufficient")
            except Exception as e:
                print("Account age filter error:", e)

# !noavatarfilter command
@bot.command(name="noavatarfilter")
async def noavatarfilter_command(ctx, state: str, mode: str = None, duration: int = None):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    global no_avatar_filter_enabled, no_avatar_action, no_avatar_timeout_duration
    state = state.lower()
    if state == "on":
        no_avatar_filter_enabled = True
        if mode is not None:
            mode = mode.lower()
            if mode not in ["ban", "kick", "timeout"]:
                await ctx.send("Please enter a valid mode: ban, kick or timeout")
                return
            no_avatar_action = mode
            if mode == "timeout":
                if duration is None:
                    await ctx.send("In timeout mode, please specify a duration (in minutes).")
                    return
                no_avatar_timeout_duration = duration
        save_settings()
        await ctx.send(f"No-avatar filter enabled. Mode: {no_avatar_action}" +
                       (f", Timeout: {no_avatar_timeout_duration} minutes" if no_avatar_action == "timeout" else ""))
    elif state == "off":
        no_avatar_filter_enabled = False
        save_settings()
        await ctx.send("No-avatar filter disabled.")
    else:
        await ctx.send("Please type 'on' or 'off'.")

# !accountagefilter command
@bot.command(name="accountagefilter")
async def accountagefilter_command(ctx, state: str, min_age: int = None, mode: str = None, duration: int = None):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    global account_age_filter_enabled, account_age_min_days, account_age_action, account_age_timeout_duration
    state = state.lower()
    if state == "off":
        account_age_filter_enabled = False
        save_settings()
        await ctx.send("Account age filter disabled.")
        return
    elif state == "on":
        if min_age is None or mode is None:
            await ctx.send("Please specify the minimum account age (in days) and a mode. Example: `!accountagefilter on 7 timeout 60`")
            return
        account_age_filter_enabled = True
        account_age_min_days = min_age
        mode = mode.lower()
        if mode not in ["ban", "kick", "timeout"]:
            await ctx.send("Please enter a valid mode: ban, kick or timeout")
            return
        account_age_action = mode
        if mode == "timeout":
            if duration is None:
                await ctx.send("In timeout mode, please specify a duration (in minutes).")
                return
            account_age_timeout_duration = duration
            save_settings()
            await ctx.send(f"Account age filter enabled: Accounts younger than {min_age} days will be timed out for {duration} minutes.")
        else:
            save_settings()
            await ctx.send(f"Account age filter enabled: Accounts younger than {min_age} days will be {mode}ned.")
    else:
        await ctx.send("Please type 'on' or 'off'.")

# !samenicknamefilter command
@bot.command(name="samenicknamefilter")
async def samenicknamefilter_command(ctx, event_name: str, state: str, limit: int = None):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    state = state.lower()
    if state == "off":
        event_nickname_limit.pop(event_name, None)
        save_settings()
        await ctx.send(f"Same nickname filter for {event_name} disabled. Users can now enter unlimited entries.")
    elif state == "on":
        if limit is None:
            await ctx.send("Please provide a limit value. Example: `!samenicknamefilter Tournament2025 on 1`")
            return
        event_nickname_limit[event_name] = limit
        save_settings()
        await ctx.send(f"Same nickname filter enabled for {event_name}. Each user can enter {limit} times.")
    else:
        await ctx.send("Please type 'on' or 'off'.")

# ---------------- Security Commands ----------------
@bot.command(name="securityauthorizedadd")
async def securityauthorizedadd(ctx, identifier: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    try:
        id_val = int(identifier.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid user or role ID.")
        return
    security_authorized_ids.add(id_val)
    save_settings()
    await ctx.send(f"{identifier} is now authorized for security commands.")

@bot.command(name="securityauthorizedremove")
async def securityauthorizedremove(ctx, identifier: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    try:
        id_val = int(identifier.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid user or role ID.")
        return
    if id_val in security_authorized_ids:
        security_authorized_ids.remove(id_val)
        save_settings()
        await ctx.send(f"{identifier} has been removed from the security authorized list.")
    else:
        await ctx.send("The specified ID was not found in the security authorized list.")

@bot.command(name="securitysettings")
async def securitysettings(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    settings_text = "**Security Settings**\n\n"
    
    noavatar_status = "Enabled" if no_avatar_filter_enabled else "Disabled"
    settings_text += f"**No-Avatar Filter**\nStatus: {noavatar_status}\nAction: {no_avatar_action}\n"
    if no_avatar_action == "timeout" and no_avatar_filter_enabled:
        settings_text += f"Timeout: {no_avatar_timeout_duration} minutes\n"
    settings_text += "\n"
    
    accountage_status = "Enabled" if account_age_filter_enabled else "Disabled"
    settings_text += f"**Account Age Filter**\nStatus: {accountage_status}\n"
    if account_age_filter_enabled:
        settings_text += f"Minimum Account Age: {account_age_min_days} days\n"
        settings_text += f"Action: {account_age_action}\n"
        if account_age_action == "timeout":
            settings_text += f"Timeout: {account_age_timeout_duration} minutes\n"
    settings_text += "\n"
    
    if security_authorized_ids:
        ids_str = ", ".join(str(i) for i in security_authorized_ids)
    else:
        ids_str = "No authorized IDs added"
    settings_text += f"**Security Authorized IDs**\n{ids_str}\n"
    
    await ctx.send(settings_text)

@bot.command(name="savesettings")
async def savesettings_command(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    save_settings()
    save_security_settings()
    await ctx.send("✅ All settings have been manually saved to bot_settings.json and security_settings.json")

@bot.command(name="loadsettings")
async def loadsettings_command(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    load_settings()
    load_security_settings()
    await ctx.send("✅ All settings have been reloaded from bot_settings.json and security_settings.json")

@bot.command(name="securityhelp")
async def securityhelp(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    help_text = (
        "**Security Commands Help Menu**\n\n"
        "1. **!noavatarfilter on [mode] [duration] / off**\n"
        "   - Description: Checks new members for an avatar. Mode options: `ban`, `kick`, `timeout`.\n"
        "   - Example: `!noavatarfilter on timeout 60` → Applies a 60-minute timeout to users without an avatar.\n\n"
        "2. **!accountagefilter on <min_days> <mode> [duration] / off**\n"
        "   - Description: Checks new members for minimum account age. Mode options: `ban`, `kick`, `timeout`.\n"
        "   - Example: `!accountagefilter on 7 timeout 60` → Applies a 60-minute timeout to accounts younger than 7 days.\n\n"
        "3. **!addwhitelistuser <user_id|@user>**\n"
        "   - Description: Adds a user to the whitelist (bypasses noavatar and account age filters).\n"
        "   - Example: `!addwhitelistuser @user` or `!addwhitelistuser 123456789`\n\n"
        "4. **!removewhitelistuser <user_id|@user>**\n"
        "   - Description: Removes a user from the whitelist.\n\n"
        "5. **!whitelistusers**\n"
        "   - Description: Lists all whitelisted users.\n\n"
        "6. **!securityauthorizedadd <id>**\n"
        "   - Description: Authorizes the specified user or role ID for security commands.\n\n"
        "7. **!securityauthorizedremove <id>**\n"
        "   - Description: Removes the specified user or role ID from the security authorized list.\n\n"
        "8. **!securitysettings**\n"
        "   - Description: Displays current security settings (filter statuses, actions, timeout durations, etc.).\n\n"
        "9. **!securityaudit**\n"
        "   - Description: Shows security audit log with recent security actions.\n\n"
        "10. **!regex <regexsettingsname> <regex>**\n"
        "   - Description: Defines/updates a regex rule with the given name. Supports `/pattern/flags` or `pattern --flags imsx`.\n\n"
        "11. **!setregexsettings <regexsettingsname> <channels>**\n"
        "   - Description: Assigns which channels the regex rule applies to. You can specify multiple channels by ID or #mention.\n"
        "   - Also supported: `!setregexsettings <name> allchannel notchannel <channels_to_exclude>` → apply to all text channels except the ones listed.\n\n"
        "12. **!setregexexempt <regexsettingsname> users|roles <targets>**\n"
        "   - Description: Sets users or roles exempt from the rule.\n\n"
        "13. **!regexsettings [regexsettingsname]**\n"
        "   - Description: Shows active regex rules and their details (channels and exemptions). Provide a name to see only that rule.\n\n"
        "14. **!delregexsettings <regexsettingsname>**\n"
        "   - Description: Deletes the specified regex setting from this server.\n\n"
        "15. **!spamrule** - Two modes available:\n"
        "   **Mod Actions:** `mod warn` (DM only), `mod delete` (delete msg), `mod warnanddelete` (both)\n\n"
        "   **Time Format:** `s`=saniye, `min`=dakika, `h`=saat, `d`=gun, `m`=ay(30 gun)\n"
        "   Examples: `30s`, `5min`, `1h`, `24h`, `7d`, `30d`, `1m`, `12m`\n\n"
        "   **Channel Options:** (after modlogchannel)\n"
        "   - `channels allchannel` - applies to all channels\n"
        "   - `channels #ch1 #ch2` - applies only to these channels\n"
        "   - `channels allchannel notchannel #ch1 #ch2` - all channels except these\n\n"
        "   **Role Options:** (mention or role ID)\n"
        "   - `roles allroles` - applies to all users (default)\n"
        "   - `roles @role1 @role2` or `roles 123456 789012` - only these roles\n"
        "   - `roles allroles exemptroles @mod 123456` - all except these roles\n\n"
        "   **Similarity Mode:**\n"
        "   `!spamrule <name> [mod action] characters>X %Y <duration> message>Z dm \"text\" modlogchannel #ch [channels ...]`\n"
        "   - Detects similar messages based on character/token similarity.\n"
        "   - Example: `!spamrule test mod warn characters>30 %80 24h message>3 dm \"Stop\" modlogchannel #alerts channels allchannel`\n\n"
        "   **Regex Mode:**\n"
        "   `!spamrule <name> [mod action] regex \"pattern\" <duration> message>Z dm \"text\" modlogchannel #ch [channels ...]`\n"
        "   - Detects messages matching a regex pattern.\n"
        "   - Example: `!spamrule linkspam mod warn regex \"https?://\\S+\" 1h message>2 dm \"Links!\" modlogchannel #alerts channels #general #chat`\n"
        "   - Example: `!spamrule invitespam mod delete regex \"discord\\.gg/\\S+\" 24h message>3 dm \"No spam\" modlogchannel #mod-log channels allchannel notchannel #bot`\n\n"
        "16. **!removespamrule <rulename>**\n"
        "   - Description: Removes a spam detection rule.\n\n"
        "17. **!spamrules [rulename]**\n"
        "   - Description: Lists all spam rules or shows details of a specific rule.\n\n"
        "18. **!setverifyrole <role_id|@role>**\n"
        "   - Description: Sets the role to be assigned after successful CAPTCHA verification.\n"
        "   - Example: `!setverifyrole @Verified` → Sets the Verified role as the verification reward.\n\n"
        "19. **!sendverifypanel [#channel|channel_id]**\n"
        "   - Description: Sends a verification panel with CAPTCHA button to the specified channel (or current channel).\n"
        "   - Example: `!sendverifypanel #verification` → Sends verification panel to the verification channel.\n\n"
        "20. **!setverifypaneltext <title|description|image> <text|url>**\n"
        "   - Description: Customizes the verification panel title, description text, or image.\n"
        "   - Examples: `!setverifypaneltext title Welcome to Our Server` → Changes panel title.\n"
        "   - `!setverifypaneltext image https://example.com/logo.png` → Adds panel image.\n\n"
        "21. **!showverifypaneltext**\n"
        "   - Description: Shows the current verification panel text settings.\n\n"
        "22. **!resetverifypaneltext**\n"
        "   - Description: Resets verification panel text to default values.\n\n"
        "23. **!savesettings**\n"
        "   - Description: Manually saves all bot settings to JSON file.\n\n"
        "24. **!loadsettings**\n"
        "   - Description: Reloads all bot settings from JSON file.\n\n"
        "25. **!savesecurity**\n"
        "   - Description: Manually saves security settings to security_settings.json.\n\n"
        "26. **!securityhelp**\n"
        "   - Description: Shows this help menu.\n"
    )
    # Split into chunks to respect Discord 2000-char message limit
    parts = []
    buffer = []
    current_len = 0
    for para in help_text.split("\n\n"):
        block = para + "\n\n"
        if current_len + len(block) > 1900:
            if buffer:
                parts.append("".join(buffer).rstrip())
                buffer = [block]
                current_len = len(block)
            else:
                # Hard-split if single block is too long
                for i in range(0, len(block), 1900):
                    parts.append(block[i:i+1900])
                buffer = []
                current_len = 0
        else:
            buffer.append(block)
            current_len += len(block)
    if buffer:
        parts.append("".join(buffer).rstrip())
    for part in parts:
        await ctx.send(part)

@bot.command(name="savesecurity")
async def savesecurity(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    # Security: Rate limiting
    if await _handle_security_rate_limit(ctx, "savesecurity"):
        return
    
    # Manuel kaydetme
    success = save_security_settings()
    
    if success:
        stats: List[str] = []
        if no_avatar_filter_enabled or account_age_filter_enabled:
            filters = []
            if no_avatar_filter_enabled:
                filters.append("No-Avatar")
            if account_age_filter_enabled:
                filters.append("Account Age")
            stats.append(f"**Active Filters:** {', '.join(filters)}")

        if security_authorized_ids:
            stats.append(f"**Authorized IDs:** {len(security_authorized_ids)}")

        if captcha_verify_role_id:
            stats.append("**Captcha Role:** Set")

        if captcha_panel_texts:
            stats.append(f"**Panel Texts:** {len(captcha_panel_texts)} guilds")

        if regex_settings_by_guild:
            total_rules = sum(len(rules) for rules in regex_settings_by_guild.values())
            stats.append(f"**Regex Rules:** {total_rules} rules in {len(regex_settings_by_guild)} guilds")

        if verify_button_usage:
            stats.append(f"**Verify Usage:** {len(verify_button_usage)} users tracked")

        message_lines = [
            "✅ Security Settings Saved",
            "",
            "All security settings have been successfully saved to file.",
        ]

        if stats:
            message_lines.extend(["", "📊 Saved Settings Summary"])
            message_lines.extend(stats)

        message_lines.extend(["", f"📁 File Location: `{SECURITY_SETTINGS_FILE}`"])

        await _send_long_message(ctx.send, "\n".join(message_lines))
    else:
        failure_message = (
            "❌ Save Failed\n\n"
            "An error occurred while saving security settings. Check console for details."
        )
        await _send_long_message(ctx.send, failure_message)

@bot.command(name="securityaudit")
async def securityaudit(ctx, limit: int = 10):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    if limit < 1 or limit > 50:
        await ctx.send("Limit must be between 1 and 50.")
        return
    
    if not security_audit_log:
        await ctx.send("No security audit entries found.")
        return
    
    # Get last N entries
    recent_entries = security_audit_log[-limit:]
    
    lines = [
        "🔍 **Security Audit Log**",
        f"Showing last {len(recent_entries)} entries",
        "",
    ]

    for i, entry in enumerate(reversed(recent_entries), 1):
        action_emoji = "➕" if "ADD" in entry["action"] else "➖"
        timestamp = entry["timestamp"][:19].replace("T", " ")  # Format: YYYY-MM-DD HH:MM:SS
        lines.append(f"{action_emoji} #{i} - {entry['action']}")
        lines.append(f"Executor: {entry['executor']}")
        lines.append(f"Target: {entry['target']}")
        lines.append(f"Time: {timestamp}")
        lines.append("")

    await _send_long_message(ctx.send, "\n".join(line for line in lines).rstrip())

# !addwhitelistuser command
@bot.command(name="addwhitelistuser")
async def addwhitelistuser_command(ctx, user_id: str):
    """Add a user to the whitelist (bypasses noavatar and account age filters)"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    # Security: Rate limiting
    if await _handle_security_rate_limit(ctx, "addwhitelistuser"):
        return
    
    # Parse user ID
    try:
        # Strip mention formatting if present
        clean_id = user_id.strip("<@!>")
        uid = int(clean_id)
    except ValueError:
        await ctx.send("❌ Please provide a valid user ID or mention. Example: `!addwhitelistuser @user` or `!addwhitelistuser 123456789`")
        return
    
    # Check if already in whitelist
    if uid in security_whitelist_users:
        await ctx.send("⚠️ This user is already in the whitelist.")
        return
    
    # Add to whitelist
    security_whitelist_users.add(uid)
    
    # Save settings
    save_security_settings()
    
    # Get user info for confirmation
    try:
        user = await bot.fetch_user(uid)
        user_display = f"{user.name} ({user.id})"
    except:
        user_display = f"User ID: {uid}"
    
    await ctx.send(f"✅ User **{user_display}** has been added to the whitelist.\n"
                   f"📊 This user will bypass noavatar and account age filters.\n"
                   f"Total whitelisted users: {len(security_whitelist_users)}")
    print(f"[SECURITY] User {uid} added to whitelist by {ctx.author.name} ({ctx.author.id})")

# !removewhitelistuser command
@bot.command(name="removewhitelistuser")
async def removewhitelistuser_command(ctx, user_id: str):
    """Remove a user from the whitelist"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    # Security: Rate limiting
    if await _handle_security_rate_limit(ctx, "removewhitelistuser"):
        return
    
    # Parse user ID
    try:
        # Strip mention formatting if present
        clean_id = user_id.strip("<@!>")
        uid = int(clean_id)
    except ValueError:
        await ctx.send("❌ Please provide a valid user ID or mention. Example: `!removewhitelistuser @user` or `!removewhitelistuser 123456789`")
        return
    
    # Check if in whitelist
    if uid not in security_whitelist_users:
        await ctx.send("⚠️ This user is not in the whitelist.")
        return
    
    # Remove from whitelist
    security_whitelist_users.remove(uid)
    
    # Save settings
    save_security_settings()
    
    # Get user info for confirmation
    try:
        user = await bot.fetch_user(uid)
        user_display = f"{user.name} ({user.id})"
    except:
        user_display = f"User ID: {uid}"
    
    await ctx.send(f"✅ User **{user_display}** has been removed from the whitelist.\n"
                   f"📊 This user will now be subject to security filters.\n"
                   f"Total whitelisted users: {len(security_whitelist_users)}")
    print(f"[SECURITY] User {uid} removed from whitelist by {ctx.author.name} ({ctx.author.id})")

# !whitelistusers command
@bot.command(name="whitelistusers")
async def whitelistusers_command(ctx):
    """List all whitelisted users"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    if not security_whitelist_users:
        await ctx.send("📋 **Whitelist is empty**\nNo users are currently whitelisted.")
        return
    
    lines = [
        "📋 **Whitelisted Users**",
        f"Total: {len(security_whitelist_users)}",
        "",
        "These users bypass noavatar and account age filters:",
        ""
    ]
    
    # Fetch user info for each whitelisted user (no mentions)
    for uid in sorted(security_whitelist_users):
        try:
            user = await bot.fetch_user(uid)
            username = getattr(user, "global_name", None) or getattr(user, "name", str(uid))
            user_info = f"• {username} (ID: {uid})"
        except Exception:
            user_info = f"• Unknown User (ID: {uid})"
        lines.append(user_info)
    
    await _send_long_message(ctx.send, "\n".join(lines))

# ============== SPAM RULE COMMANDS ==============

@bot.command(name="spamrule")
async def spamrule(ctx, rulename: str, *, rule_spec: str = ""):
    """Create or update a spam rule with similarity detection"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return

    if ctx.guild is None:
        await ctx.send("This command can only be used inside a server.")
        return

    if await _handle_security_rate_limit(ctx, "spamrule"):
        return

    rule_spec = (rule_spec or "").strip()
    if not rule_spec:
        await ctx.send(
            "Please provide rule details. Example: `!spamrule test characters>30 %80 24h message>3 dm \"Your DM\" modlogchannel #alerts`."
        )
        return

    try:
        parts = shlex.split(rule_spec)
    except ValueError as exc:
        await ctx.send(f"Unable to parse command parameters: {exc}")
        return

    mod_action: str | None = None
    if parts and parts[0].lower() == "mod":
        parts.pop(0)
        if not parts:
            await ctx.send("Specify a moderation action after `mod` (warn, delete, warnanddelete).")
            return
        mod_action_token = parts.pop(0).lower()
        if mod_action_token not in {"warn", "delete", "warnanddelete"}:
            await ctx.send("Unknown moderation action. Use `mod warn`, `mod delete`, or `mod warnanddelete`.")
            return
        mod_action = mod_action_token

    # Check if using regex mode or similarity mode
    regex_pattern_str: str | None = None
    compiled_regex: re.Pattern | None = None

    if parts and parts[0].lower() == "regex":
        # REGEX MODE
        parts.pop(0)  # Remove "regex" keyword
        if not parts:
            await ctx.send("Provide a regex pattern after `regex` keyword (wrap in quotes if it has spaces).")
            return
        regex_pattern_str = parts.pop(0)
        # Strip invisible/zero-width Unicode characters that Discord might add
        regex_pattern_str = ''.join(c for c in regex_pattern_str if c.isprintable() or c in '\t\n\r')
        try:
            compiled_regex = re.compile(regex_pattern_str, re.IGNORECASE)
        except re.error as exc:
            await ctx.send(f"Invalid regex pattern: {exc}")
            return
        # In regex mode, min_length and similarity are not used
        min_length = 0
        similarity_threshold = 0.0

        # Regex mode requires: <duration>, message>, optional dm, modlogchannel
        if len(parts) < 3:
            await ctx.send(
                "Invalid format for regex mode. Expected `regex \"pattern\" <duration> message>N dm \"text\" modlogchannel #channel`."
            )
            return
    else:
        # SIMILARITY MODE (original behavior)
        if len(parts) < 5:
            await ctx.send(
                "Invalid format. Expected `characters>`, `%`, `<duration>`, `message>`, optional `dm`, then your DM text and channels."
            )
            return

        length_token = parts.pop(0)
        length_match = re.fullmatch(r"characters\s*>\s*(\d+)", length_token, flags=re.IGNORECASE)
        if not length_match:
            await ctx.send("Specify minimum characters like `characters>30`.")
            return
        min_length = int(length_match.group(1))

        similarity_token = parts.pop(0)
        similarity_match = None
        for pattern in (r"%\s*(\d+(?:\.\d+)?)", r"(\d+(?:\.\d+)?)%", r"similarity\s*>\s*(\d+(?:\.\d+)?)"):
            similarity_match = re.fullmatch(pattern, similarity_token, flags=re.IGNORECASE)
            if similarity_match:
                break
        if not similarity_match:
            await ctx.send("Specify similarity like `%80` or `80%`.")
            return
        similarity_value = float(similarity_match.group(1))
        similarity_threshold = max(0.0, min(similarity_value / 100.0, 1.0))

    duration_token = parts.pop(0).lower()
    duration_display = duration_token
    time_window = SPAM_RULE_PREDEFINED_WINDOWS.get(duration_token)
    if time_window is not None:
        duration_match = re.fullmatch(r"(\d+)([a-z]+)", duration_token)
    else:
        duration_match = re.fullmatch(r"(\d+)(s|min|h|d|m)", duration_token)
        if not duration_match:
            await ctx.send("Specify time window like `24h`, `7d`, `5min`, or `1m` (month).")
            return
        window_value = int(duration_match.group(1))
        window_unit = duration_match.group(2)
        unit_multipliers = {"s": 1, "min": 60, "h": 3600, "d": 86400, "m": 2592000}
        time_window = window_value * unit_multipliers[window_unit]
        duration_display = f"{window_value}{window_unit}"

    message_token = parts.pop(0)
    message_match = re.fullmatch(r"messages?\s*>\s*(\d+)", message_token, flags=re.IGNORECASE)
    if not message_match:
        await ctx.send("Specify message threshold like `message>3` or `messages>3`.")
        return
    message_count = int(message_match.group(1))

    KEYWORDS = {"modlogchannel", "channels", "allchannel", "notchannel", "nonreply", "roles", "allroles", "exemptroles"}

    def _is_keyword(token: str) -> bool:
        lowered = token.lower()
        return lowered in KEYWORDS or lowered.startswith("nonreply")

    dm_message = ""
    if parts and parts[0].lower() == "dm":
        parts.pop(0)

        dm_tokens: list[str] = []
        while parts and not _is_keyword(parts[0]):
            dm_tokens.append(parts.pop(0))

        if not dm_tokens:
            await ctx.send("Provide the message to send via DM after the `dm` keyword (wrap in quotes if it has spaces).")
            return

        dm_message = " ".join(dm_tokens).strip()
        if not dm_message:
            await ctx.send("Provide the message to send via DM after the `dm` keyword (wrap in quotes if it has spaces).")
            return
    else:
        if mod_action in (None, "warn", "warnanddelete"):
            await ctx.send("Please include `dm` followed by the message to send.")
            return

    def _resolve_channel(token: str) -> discord.abc.GuildChannel | None:
        raw = token.strip()
        if raw.startswith("<#") and raw.endswith(">"):
            try:
                channel_id_inner = int(raw[2:-1])
            except ValueError:
                return None
            return ctx.guild.get_channel(channel_id_inner)
        if raw.startswith("#"):
            name_inner = raw[1:]
            return discord.utils.get(ctx.guild.text_channels, name=name_inner)
        try:
            channel_id_inner = int(raw)
        except ValueError:
            return None
        return ctx.guild.get_channel(channel_id_inner)

    def _resolve_role(token: str) -> discord.Role | None:
        raw = token.strip()
        if raw.startswith("<@&") and raw.endswith(">"):
            try:
                role_id_inner = int(raw[3:-1])
            except ValueError:
                return None
            return ctx.guild.get_role(role_id_inner)
        if raw.startswith("@"):
            name_inner = raw[1:]
            return discord.utils.get(ctx.guild.roles, name=name_inner)
        try:
            role_id_inner = int(raw)
        except ValueError:
            return None
        return ctx.guild.get_role(role_id_inner)

    notify_channel: discord.TextChannel | None = None
    monitored_channels: set[int] = set()
    excluded_channels: set[int] = set()
    targeted_roles: set[int] = set()
    exempted_roles: set[int] = set()
    nonreply_only = False

    while parts:
        token = parts.pop(0)
        lowered = token.lower()
        if lowered == "modlogchannel":
            if not parts:
                await ctx.send("Please provide a channel after `modlogchannel`.")
                return
            channel_token = parts.pop(0)
            channel = _resolve_channel(channel_token)
            if not isinstance(channel, discord.TextChannel):
                await ctx.send("Please mention a valid text channel after `modlogchannel`.")
                return
            notify_channel = channel
            continue
        if lowered == "channels":
            if not parts:
                await ctx.send("Provide channel option after `channels` (allchannel, #ch1 #ch2, or allchannel notchannel #ch).")
                return

            next_token = parts[0].lower()
            if next_token == "allchannel":
                parts.pop(0)  # Remove "allchannel"
                # Check if followed by "notchannel" for exclusion
                if parts and parts[0].lower() == "notchannel":
                    parts.pop(0)  # Remove "notchannel"
                    if not parts:
                        await ctx.send("Provide at least one channel after `notchannel`.")
                        return
                    excl_found = False
                    while parts and not _is_keyword(parts[0]):
                        channel_token = parts.pop(0)
                        channel = _resolve_channel(channel_token)
                        if not isinstance(channel, discord.TextChannel):
                            await ctx.send("`notchannel` must be followed by valid text channel mentions.")
                            return
                        excluded_channels.add(channel.id)
                        excl_found = True
                    if not excl_found:
                        await ctx.send("Provide at least one channel after `notchannel`.")
                        return
                # allchannel alone = all channels (monitored_channels stays empty)
            else:
                # Specific channels: channels #ch1 #ch2
                spec_found = False
                while parts and not _is_keyword(parts[0]):
                    channel_token = parts.pop(0)
                    channel = _resolve_channel(channel_token)
                    if not isinstance(channel, discord.TextChannel):
                        await ctx.send("`channels` must be followed by valid text channel mentions.")
                        return
                    monitored_channels.add(channel.id)
                    spec_found = True
                if not spec_found:
                    await ctx.send("Provide at least one channel after `channels`.")
                    return
            continue
        if lowered == "roles":
            if not parts:
                await ctx.send("Provide role option after `roles` (allroles, @role1 @role2, or allroles exemptroles @role).")
                return

            next_token = parts[0].lower()
            if next_token == "allroles":
                parts.pop(0)  # Remove "allroles"
                # Check if followed by "exemptroles" for exclusion
                if parts and parts[0].lower() == "exemptroles":
                    parts.pop(0)  # Remove "exemptroles"
                    if not parts:
                        await ctx.send("Provide at least one role after `exemptroles`.")
                        return
                    exempt_found = False
                    while parts and not _is_keyword(parts[0]):
                        role_token = parts.pop(0)
                        role = _resolve_role(role_token)
                        if role is None:
                            await ctx.send(f"`exemptroles` must be followed by valid role mentions. Could not resolve `{role_token}`.")
                            return
                        exempted_roles.add(role.id)
                        exempt_found = True
                    if not exempt_found:
                        await ctx.send("Provide at least one role after `exemptroles`.")
                        return
                # allroles alone = all roles (targeted_roles stays empty)
            else:
                # Specific roles: roles @role1 @role2
                spec_found = False
                while parts and not _is_keyword(parts[0]):
                    role_token = parts.pop(0)
                    role = _resolve_role(role_token)
                    if role is None:
                        await ctx.send(f"`roles` must be followed by valid role mentions. Could not resolve `{role_token}`.")
                        return
                    targeted_roles.add(role.id)
                    spec_found = True
                if not spec_found:
                    await ctx.send("Provide at least one role after `roles`.")
                    return
            continue
        if lowered.startswith("nonreply"):
            state_token = None

            # Allow inline forms like nonreply=on or nonreply:on
            inline = lowered[len("nonreply"):].lstrip(" =:")
            if inline:
                state_token = inline

            if state_token is None and parts and not _is_keyword(parts[0]):
                state_token = parts.pop(0).lower()

            if state_token is None:
                nonreply_only = True
                continue
            if state_token in ("on", "true", "yes", "1", "enable", "enabled"):
                nonreply_only = True
                continue
            if state_token in ("off", "false", "no", "0", "disable", "disabled"):
                nonreply_only = False
                continue
            await ctx.send("Use `nonreply on` or `nonreply off` (inline forms like `nonreply=on` also work).")
            return

        channel = _resolve_channel(token)
        if isinstance(channel, discord.TextChannel):
            if notify_channel is None:
                notify_channel = channel
            else:
                monitored_channels.add(channel.id)
            continue

        await ctx.send(f"Unrecognized token `{token}` in command arguments.")
        return

    if notify_channel is None:
        await ctx.send(
            "Please specify the moderation alert channel using `modlogchannel #channel` (or mention a channel directly)."
        )
        return

    monitored_channels = {cid for cid in monitored_channels if ctx.guild.get_channel(cid)}

    guild_id = ctx.guild.id
    name_key = rulename.strip().lower()
    label = rulename.strip() or name_key

    guild_rules = spam_rules_by_guild.setdefault(guild_id, {})

    _reset_spam_history_for_rule(guild_id, name_key)
    await remove_spam_violation_stats_for_rule(guild_id, name_key)

    guild_rules[name_key] = {
        "label": label,
        "min_length": max(0, min_length),
        "similarity_threshold": similarity_threshold,
        "time_window": max(0, time_window),
        "message_count": max(0, message_count),
        "dm_message": dm_message,
        "notify_channel_id": notify_channel.id,
        "channels": monitored_channels,
        "excluded_channels": excluded_channels,
        "targeted_roles": targeted_roles,
        "exempted_roles": exempted_roles,
        "nonreply_only": nonreply_only,
        "mod_action": mod_action,
        "regex_pattern": regex_pattern_str,  # None for similarity mode, pattern string for regex mode
    }

    save_security_settings()

    details = [f"Spam rule `{label}` saved."]
    if regex_pattern_str:
        details.append(f"- Mode: Regex")
        details.append(f"- Pattern: `{regex_pattern_str}`")
    else:
        details.append(f"- Mode: Similarity")
        details.append(f"- Min characters: {min_length}")
        details.append(f"- Similarity: {similarity_threshold * 100:.0f}%")
    details.extend([
        f"- Window: {duration_display}",
        f"- Message count: {message_count}",
        f"- Notify: {notify_channel.mention}",
    ])
    if mod_action:
        action_description = {
            "warn": "Warn via DM",
            "delete": "Delete matching messages",
            "warnanddelete": "Warn via DM and delete messages",
        }.get(mod_action, mod_action)
        details.append(f"- Action: {action_description}")
    else:
        details.append("- Action: Warn via DM (default)")
    if monitored_channels:
        channel_mentions = ", ".join(f"<#{cid}>" for cid in monitored_channels)
        details.append(f"- Monitored channels: {channel_mentions}")
    elif excluded_channels:
        excl_mentions = ", ".join(f"<#{cid}>" for cid in excluded_channels)
        details.append(f"- Monitored channels: all except {excl_mentions}")
    else:
        details.append("- Monitored channels: all text channels")
    if targeted_roles:
        role_names = ", ".join(ctx.guild.get_role(rid).name if ctx.guild.get_role(rid) else str(rid) for rid in targeted_roles)
        details.append(f"- Targeted roles: {role_names}")
    elif exempted_roles:
        exempt_names = ", ".join(ctx.guild.get_role(rid).name if ctx.guild.get_role(rid) else str(rid) for rid in exempted_roles)
        details.append(f"- Targeted roles: all except {exempt_names}")
    else:
        details.append("- Targeted roles: all users")
    details.append(f"- Count only non-replies: {'Yes' if nonreply_only else 'No'}")

    await ctx.send("\n".join(details))

@bot.command(name="removespamrule")
async def removespamrule(ctx, rulename: str):
    """Remove a previously configured spam rule"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return

    if ctx.guild is None:
        await ctx.send("This command can only be used inside a server.")
        return

    if await _handle_security_rate_limit(ctx, "removespamrule"):
        return

    guild_id = ctx.guild.id
    guild_rules = spam_rules_by_guild.get(guild_id)
    if not guild_rules:
        await ctx.send("No spam rules are configured for this server.")
        return

    name_key = rulename.strip().lower()
    if name_key not in guild_rules:
        await ctx.send("No spam rule found with that name.")
        return

    removed_rule = guild_rules.pop(name_key, None)
    if not guild_rules:
        try:
            del spam_rules_by_guild[guild_id]
        except KeyError:
            pass

    # Clean trigger log entries for this rule in this guild
    keys_to_delete = [key for key in spam_rule_trigger_log if key[0] == guild_id and key[2] == name_key]
    for key in keys_to_delete:
        del spam_rule_trigger_log[key]

    # Save settings after removal
    save_security_settings()

    await remove_spam_violation_stats_for_rule(guild_id, name_key)

    label = removed_rule.get("label") if removed_rule else rulename
    await ctx.send(f"Spam rule `{label}` has been removed.")


@bot.command(name="spamrules")
async def spamrules(ctx):
    """List configured spam rules"""
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return

    if ctx.guild is None:
        await ctx.send("This command can only be used inside a server.")
        return

    guild_rules = spam_rules_by_guild.get(ctx.guild.id)
    if not guild_rules:
        await ctx.send("No spam rules are currently configured for this server.")
        return

    def _format_window(seconds: int) -> str:
        seconds = max(0, int(seconds))
        if seconds == 0:
            return "0s"
        if seconds % 86400 == 0:
            return f"{seconds // 86400}d"
        if seconds % 3600 == 0:
            return f"{seconds // 3600}h"
        if seconds % 60 == 0:
            return f"{seconds // 60}m"
        return f"{seconds}s"

    lines: list[str] = ["🧩 **Configured Spam Rules**"]
    for name_key, rule in sorted(guild_rules.items()):
        label = rule.get("label", name_key)
        min_length = rule.get("min_length", 0)
        similarity = float(rule.get("similarity_threshold", 0.0)) * 100
        regex_pattern = rule.get("regex_pattern")
        message_count = rule.get("message_count", 0)
        window = _format_window(rule.get("time_window", 0))
        nonreply_only = _coerce_bool(rule.get("nonreply_only", False))
        mod_action = (rule.get("mod_action") or "").lower()

        notify_channel_id = rule.get("notify_channel_id")
        notify_channel = None
        if notify_channel_id:
            notify_channel = ctx.guild.get_channel(int(notify_channel_id))

        channels = rule.get("channels", set()) or set()
        excluded_channels = rule.get("excluded_channels", set()) or set()
        targeted_roles = rule.get("targeted_roles", set()) or set()
        exempted_roles = rule.get("exempted_roles", set()) or set()
        if channels:
            channel_mentions = []
            for channel_id in sorted(channels):
                channel_obj = ctx.guild.get_channel(int(channel_id))
                if isinstance(channel_obj, discord.TextChannel):
                    channel_mentions.append(channel_obj.mention)
                else:
                    channel_mentions.append(f"`{channel_id}`")
            channels_text = ", ".join(channel_mentions)
        elif excluded_channels:
            excl_mentions = []
            for channel_id in sorted(excluded_channels):
                channel_obj = ctx.guild.get_channel(int(channel_id))
                if isinstance(channel_obj, discord.TextChannel):
                    excl_mentions.append(channel_obj.mention)
                else:
                    excl_mentions.append(f"`{channel_id}`")
            channels_text = f"All except {', '.join(excl_mentions)}"
        else:
            channels_text = "All text channels"

        # Role targeting
        if targeted_roles:
            role_names = []
            for role_id in sorted(targeted_roles):
                role_obj = ctx.guild.get_role(int(role_id))
                if role_obj:
                    role_names.append(role_obj.name)
                else:
                    role_names.append(f"`{role_id}`")
            roles_text = ", ".join(role_names)
        elif exempted_roles:
            exempt_names = []
            for role_id in sorted(exempted_roles):
                role_obj = ctx.guild.get_role(int(role_id))
                if role_obj:
                    exempt_names.append(role_obj.name)
                else:
                    exempt_names.append(f"`{role_id}`")
            roles_text = f"All except {', '.join(exempt_names)}"
        else:
            roles_text = "All users"

        notify_text = (
            notify_channel.mention
            if isinstance(notify_channel, discord.TextChannel)
            else (f"`{notify_channel_id}`" if notify_channel_id else "Not set")
        )

        lines.append(f"\n**{label}** (`{name_key}`)")
        if regex_pattern:
            lines.append(f"• Mode: Regex")
            lines.append(f"• Pattern: `{regex_pattern}`")
        else:
            lines.append(f"• Mode: Similarity")
            lines.append(f"• Min characters: {min_length}")
            lines.append(f"• Similarity: {similarity:.0f}%")
        lines.extend([
            f"• Threshold: {message_count} messages in {window}",
            f"• Mod-log channel: {notify_text}",
            f"• Scope: {channels_text}",
            f"• Roles: {roles_text}",
            f"• Count only non-replies: {'Yes' if nonreply_only else 'No'}",
        ])

        if mod_action == "warn":
            action_text = "mod warn (warn via DM)"
        elif mod_action == "delete":
            action_text = "mod delete (delete matching messages)"
        elif mod_action == "warnanddelete":
            action_text = "mod warnanddelete (warn via DM and delete messages)"
        else:
            action_text = "default (warn via DM)"
        lines.append(f"• Action: {action_text}")

        dm_message = rule.get("dm_message")
        if dm_message:
            preview = dm_message if len(dm_message) <= 120 else dm_message[:117] + "..."
            lines.append(f"• DM message: {preview}")

    messages: list[str] = []
    buffer: list[str] = []
    length = 0
    for line in lines:
        addition = len(line) + 1
        if buffer and length + addition > 1900:
            messages.append("\n".join(buffer))
            buffer = [line]
            length = len(line)
        else:
            buffer.append(line)
            length += addition

    if buffer:
        messages.append("\n".join(buffer))

    for chunk in messages:
        await ctx.send(chunk)

# ---------------- Play Event Section ----------------
@bot.command(name="createplayevent")
async def createplayevent(ctx, event_name: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name in events:
        await ctx.send(f"{event_name} event already exists. Recreating with the same name will reset usage counts.")
        return
    events[event_name] = {
        "link": None,
        "password": None,  # Added password field
        "channel_id": None,
        "excel_file": f"{event_name}_play_records.xlsx",
        "limits": {}
    }
    event_nickname_counts[event_name] = {}  # Initialize same nickname counter.
    save_settings()
    await ctx.send(f"{event_name} event has been created.")

@bot.command(name="setplaylink")
async def setplaylink(ctx, event_name: str, link: str, *args):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name not in events:
        await ctx.send("Please create the event first using !createplayevent.")
        return
    
    # Set the link
    events[event_name]["link"] = link
    
    # Check for password in the arguments
    password = None
    for i, arg in enumerate(args):
        if arg.lower() == "password" and i+1 < len(args):
            password = args[i+1]
            break
    
    # If password is found, store it
    if password:
        events[event_name]["password"] = password
        save_settings()
        await ctx.send(f"Link set for {event_name} event: {link}\nPassword: {password}")
    else:
        events[event_name]["password"] = None
        save_settings()
        await ctx.send(f"Link set for {event_name} event: {link}")

@bot.command(name="setplaychannel")
async def setplaychannel(ctx, event_name: str, channel_input: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name not in events:
        await ctx.send("Please create the event first using !createplayevent.")
        return
    try:
        channel_id = int(channel_input.strip("<#>"))
    except ValueError:
        await ctx.send("Please provide a valid channel ID or channel mention.")
        return
    channel = ctx.guild.get_channel(channel_id)
    if channel is None:
        await ctx.send("No channel found with the provided ID.")
        return
    events[event_name]["channel_id"] = channel.id
    save_settings()
    await ctx.send(f"Channel set for {event_name} event: {channel.mention}")

@bot.command(name="setauthorizedrole")
async def setauthorizedrole(ctx, role_input: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    global authorized_role_id
    try:
        role_id = int(role_input.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid role ID or role mention.")
        return
    role = ctx.guild.get_role(role_id)
    if role is None:
        await ctx.send("No role found with the provided ID.")
        return
    authorized_role_id = role.id
    await ctx.send(f"Authorized role for play event commands set to: {role.mention}")

@bot.command(name="sendplay")
async def sendplay(ctx, event_name: str, channel_input: str = None):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    
    if event_name not in events:
        await ctx.send("Event not found. Please create the event first using !createplayevent.")
        return
    
    # Determine which channel to use
    target_channel = None
    if channel_input:
        try:
            channel_id = int(channel_input.strip("<#>"))
            target_channel = ctx.guild.get_channel(channel_id)
            if target_channel is None:
                await ctx.send("No channel found with the provided ID.")
                return
        except ValueError:
            await ctx.send("Please provide a valid channel ID or channel mention.")
            return
    else:
        # Use the channel set in the event if no channel is specified
        channel_id = events[event_name]["channel_id"]
        if channel_id:
            target_channel = ctx.guild.get_channel(channel_id)
        else:
            await ctx.send("No channel is set for this event. Please specify a channel.")
            return
    
    # Create a button for the event
    class PlayButton(discord.ui.View):
        def __init__(self):
            super().__init__(timeout=None)
            self.add_item(discord.ui.Button(
                style=discord.ButtonStyle.primary,
                label="Play",
                custom_id=f"play_button_{event_name}"
            ))
    
    # Send the button
    try:
        await target_channel.send(
            f"Click the button below to register for **{event_name}**:",
            view=PlayButton()
        )
        await ctx.send(f"Play button for {event_name} has been sent to {target_channel.mention}")
    except Exception as e:
        await ctx.send(f"Failed to send play button: {str(e)}")

@bot.command(name="removeplaybutton")
async def removeplaybutton(ctx, event_name: str, channel_input: str = None):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    
    if event_name not in events:
        await ctx.send("Event not found.")
        return
    
    # Determine which channel to use
    target_channel = None
    if channel_input:
        try:
            channel_id = int(channel_input.strip("<#>"))
            target_channel = ctx.guild.get_channel(channel_id)
            if target_channel is None:
                await ctx.send("No channel found with the provided ID.")
                return
        except ValueError:
            await ctx.send("Please provide a valid channel ID or channel mention.")
            return
    else:
        # Use the channel set in the event if no channel is specified
        channel_id = events[event_name]["channel_id"]
        if channel_id:
            target_channel = ctx.guild.get_channel(channel_id)
        else:
            await ctx.send("No channel is set for this event. Please specify a channel.")
            return
    
    # Try to find and delete the play button message
    try:
        async for message in target_channel.history(limit=100):
            if message.components and f"play_button_{event_name}" in str(message.components):
                await message.delete()
                await ctx.send(f"Play button for {event_name} has been removed from {target_channel.mention}")
                return
        await ctx.send(f"No play button found for {event_name} in {target_channel.mention}")
    except Exception as e:
        await ctx.send(f"Failed to remove play button: {str(e)}")

@bot.command(name="setallowedrole")
async def setallowedrole(ctx, *, roles: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    global allowed_role_ids
    role_ids = []
    for role_str in roles.split(","):
        role_str = role_str.strip()
        if role_str.startswith("<@&") and role_str.endswith(">"):
            role_str = role_str[3:-1]
        try:
            role_id = int(role_str)
            role_ids.append(role_id)
        except ValueError:
            continue
    allowed_role_ids = role_ids
    save_settings()
    await ctx.send(f"Roles allowed for button usage set to: {allowed_role_ids}")

@bot.command(name="removeallowedrole")
async def removeallowedrole(ctx, *, roles: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    global allowed_role_ids
    removed = []
    for role_str in roles.split(","):
        role_str = role_str.strip()
        if role_str.startswith("<@&") and role_str.endswith(">"):
            role_str = role_str[3:-1]
        try:
            role_id = int(role_str)
            if role_id in allowed_role_ids:
                allowed_role_ids.remove(role_id)
                removed.append(role_id)
        except ValueError:
            continue
    if removed:
        save_settings()
        await ctx.send(f"Removed roles from allowed list: {removed}")
    else:
        await ctx.send("No roles were removed from the allowed list.")

@bot.command(name="playauthorizedadd")
async def playauthorizedadd(ctx, identifier: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    try:
        id_val = int(identifier.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid user or role ID.")
        return
    play_authorized_ids.add(id_val)
    save_settings()
    await ctx.send(f"{identifier} is now authorized for play event commands.")

@bot.command(name="playauthorizedremove")
async def playauthorizedremove(ctx, identifier: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    try:
        id_val = int(identifier.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid user or role ID.")
        return
    if id_val in play_authorized_ids:
        play_authorized_ids.remove(id_val)
        save_settings()
        await ctx.send(f"{identifier} has been removed from the play authorized list.")
    else:
        await ctx.send("The specified ID was not found in the play authorized list.")

@bot.command(name="sendplaylimit")
async def sendplaylimit(ctx, event_name: str, role_input: str, limit: int):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name not in events:
        await ctx.send("Please create the event first using !createplayevent.")
        return
    try:
        role_id = int(role_input.strip("<@&>"))
    except ValueError:
        await ctx.send("Please provide a valid role ID or role mention.")
        return
    role = ctx.guild.get_role(role_id)
    if role is None:
        await ctx.send("No role found with the provided ID.")
        return
    events[event_name]["limits"][role.id] = limit
    save_settings()
    await ctx.send(f"Interaction limit for {role.mention} in {event_name} event set to {limit}.")

@bot.command(name="sendplaysettings")
async def sendplaysettings(ctx, event_name: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name not in events:
        await ctx.send("Event not found.")
        return
    event_data = events[event_name]
    settings_text = f"**{event_name} Event Settings**\n\n"
    settings_text += f"**Link:** {event_data['link'] or 'Not set'}\n"
    
    # Add password field if it exists
    if event_data.get("password"):
        settings_text += f"**Password:** {event_data['password']}\n"
    
    channel_id = event_data["channel_id"]
    channel_mention = f"<#{channel_id}>" if channel_id else "Not set"
    settings_text += f"**Channel:** {channel_mention}\n"
    settings_text += f"**Excel File:** {event_data['excel_file']}\n\n"
    
    settings_text += "**Interaction Limits:**\n"
    limits_text = ""
    for role_id, limit in event_data["limits"].items():
        role = ctx.guild.get_role(role_id)
        role_name = role.name if role else f"Unknown Role ({role_id})"
        limits_text += f"  {role_name}: {limit}\n"
    settings_text += limits_text or "  No limits set\n"
    
    settings_text += "\n"
    same_nickname_limit_text = str(event_nickname_limit.get(event_name, "No limit")) if event_name in event_nickname_limit else "No limit"
    settings_text += f"**Same Nickname Limit:** {same_nickname_limit_text}\n"
    
    await ctx.send(settings_text)

@bot.command(name="getplayexcel")
async def getplayexcel(ctx, event_name: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    excel_file_name = f"{event_name}_play_records.xlsx"
    if not os.path.exists(excel_file_name):
        await ctx.send("Excel file for the specified event not found.")
        return
    await ctx.send(file=discord.File(excel_file_name))

@bot.command(name="deletesendplay")
async def deletesendplay(ctx, event_name: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if event_name not in events:
        await ctx.send("Event not found.")
        return
    events.pop(event_name)
    event_nickname_counts.pop(event_name, None)
    event_nickname_limit.pop(event_name, None)
    save_settings()
    excel_file_name = f"{event_name}_play_records.xlsx"
    if os.path.exists(excel_file_name):
        try:
            os.remove(excel_file_name)
            await ctx.send(f"{event_name} event and its Excel file have been deleted.")
        except Exception as e:
            await ctx.send(f"Event deleted, but could not delete Excel file: {e}")
    else:
        await ctx.send(f"{event_name} event has been deleted. No Excel file was found.")

@bot.command(name="playlistid")
async def playlistid(ctx, event_name: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    excel_file_name = f"{event_name}_play_records.xlsx"
    if not os.path.exists(excel_file_name):
        await ctx.send("Excel file for the specified event not found.")
        return
    try:
        workbook = openpyxl.load_workbook(excel_file_name)
        sheet = workbook.active
    except Exception as e:
        await ctx.send(f"Error reading the Excel file: {e}")
        return
    discord_ids = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[0]  # Discord ID is in the first column
        if value is None:
            continue
        if isinstance(value, (int, float)):
            value_str = str(int(value))
        else:
            value_str = str(value)
        discord_ids.append(value_str)
    # Remove duplicate IDs while preserving order.
    unique_ids = list(dict.fromkeys(discord_ids))
    if not unique_ids:
        await ctx.send("No participants found for the event.")
        return
    chunk_size = 150
    paragraphs = []
    for i in range(0, len(unique_ids), chunk_size):
        chunk = unique_ids[i:i+chunk_size]
        # Join IDs with a space.
        paragraph = " ".join(chunk)
        paragraphs.append(paragraph)
    # Each group of 150 IDs starts on a new paragraph (separated by two newlines).
    txt_content = "\n\n".join(paragraphs)
    output_file_name = f"{event_name}_playlist.txt"
    with open(output_file_name, "w", encoding="utf-8") as f:
        f.write(txt_content)
    # Send the text file as an attachment.
    await ctx.send(file=discord.File(output_file_name))

# Updated command: !checkgameusername - Checks if game usernames match with Discord IDs
# Now supports case-insensitive matching and fuzzy matching
@bot.command(name="checkgameusername")
async def checkgameusername(ctx, first_param: str, event_name: str = None, *, usernames: str = None):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    
    # Check if the first parameter is "id"
    id_only_mode = False
    if first_param.lower() == "id":
        id_only_mode = True
        if event_name is None or usernames is None:
            await ctx.send("Usage: !checkgameusername id <event_name> <username1 username2 ...>")
            return
    else:
        # If first parameter is not "id", then it's the event_name
        if usernames is None:
            usernames = event_name
            event_name = first_param
    
    excel_file_name = f"{event_name}_play_records.xlsx"
    if not os.path.exists(excel_file_name):
        await ctx.send("Excel file for the specified event not found.")
        return
    
    try:
        workbook = openpyxl.load_workbook(excel_file_name)
        sheet = workbook.active
    except Exception as e:
        await ctx.send(f"Error reading the Excel file: {e}")
        return
    
    # Parse the usernames from the input
    username_list = [name.strip() for name in usernames.split()]
    
    # Create a dictionary to store in-game username to Discord ID mappings
    username_to_discord = {}
    
    # Read the Excel file and populate the dictionary
    for row in sheet.iter_rows(min_row=2, values_only=True):
        discord_id = str(row[0])
        in_game_username = row[2]
        if in_game_username:
            username_to_discord[in_game_username.lower()] = discord_id  # Store lowercase for case-insensitive matching
    
    # Find matches (exact, case-insensitive, and fuzzy)
    exact_matches = []
    case_insensitive_matches = []
    fuzzy_matches = []
    
    for username in username_list:
        username_lower = username.lower()
        
        # Check for exact match
        if username in username_to_discord:
            exact_matches.append((username, username_to_discord[username]))
        # Check for case-insensitive match
        elif username_lower in username_to_discord:
            case_insensitive_matches.append((username, username_to_discord[username_lower]))
        else:
            # Check for fuzzy matches (allowing 2-3 character differences)
            best_match = None
            best_similarity = 0.7  # Threshold for fuzzy matching (adjust as needed)
            
            for db_username, discord_id in username_to_discord.items():
                similarity = string_similarity(username, db_username)
                if similarity > best_similarity:
                    best_similarity = similarity
                    best_match = (username, discord_id, db_username, similarity)
            
            if best_match:
                fuzzy_matches.append(best_match)
    
    # Combine all matches
    all_matches = exact_matches + case_insensitive_matches + fuzzy_matches
    
    # Create result message
    result_message = f"Checked {len(username_list)} game usernames for event '{event_name}'.\n"
    result_message += f"Found {len(all_matches)} matching Discord IDs (Exact: {len(exact_matches)}, Case-insensitive: {len(case_insensitive_matches)}, Fuzzy: {len(fuzzy_matches)})."
    
    # Create a text file with the results
    output_file_name = f"{event_name}_username_matches.txt"
    with open(output_file_name, "w", encoding="utf-8") as f:
        f.write(f"Event: {event_name}\n")
        f.write(f"Total usernames checked: {len(username_list)}\n")
        f.write(f"Total matches found: {len(all_matches)}\n")
        f.write(f"- Exact matches: {len(exact_matches)}\n")
        f.write(f"- Case-insensitive matches: {len(case_insensitive_matches)}\n")
        f.write(f"- Fuzzy matches: {len(fuzzy_matches)}\n\n")
        
        if id_only_mode:
            # In ID-only mode, just list the Discord IDs
            f.write("Discord IDs:\n")
            for match in exact_matches:
                f.write(f"{match[1]}\n")
            for match in case_insensitive_matches:
                f.write(f"{match[1]}\n")
            for match in fuzzy_matches:
                f.write(f"{match[1]}\n")
        else:
            # In normal mode, list both username and Discord ID
            if exact_matches:
                f.write("Exact Matches (Game Username : Discord ID):\n")
                for username, discord_id in exact_matches:
                    f.write(f"{username} : {discord_id}\n")
                f.write("\n")
            
            if case_insensitive_matches:
                f.write("Case-Insensitive Matches (Game Username : Discord ID):\n")
                for username, discord_id in case_insensitive_matches:
                    f.write(f"{username} : {discord_id}\n")
                f.write("\n")
            
            if fuzzy_matches:
                f.write("Fuzzy Matches (Game Username : Discord ID : Database Username : Similarity):\n")
                for username, discord_id, db_username, similarity in fuzzy_matches:
                    f.write(f"{username} : {discord_id} : {db_username} : {similarity:.2f}\n")
    
    # Send the results
    await ctx.send(result_message, file=discord.File(output_file_name))

# Alias for checkgameusername with id parameter
@bot.command(name="checkgameusernameid")
async def checkgameusernameid(ctx, event_name: str, *, usernames: str):
    # This is just a convenience alias that calls checkgameusername with the id parameter
    await checkgameusername(ctx, "id", event_name, usernames=usernames)

@bot.command(name="allplaylist")
async def allplaylist(ctx):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    if not events:
        await ctx.send("No events found.")
        return
    event_names = "\n".join(events.keys())
    await ctx.send(f"Created events:\n{event_names}")

# Specialized command: !getusername - Optimized for Lepoker.io
@bot.command(name="getusername")
async def getusername(ctx, url: str, selector: str):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    
    await ctx.send("Extracting player names from the Lepoker.io webpage. This may take a moment...")
    
    try:
        # Extract player names using the specialized Lepoker.io function
        player_names = await extract_lepoker_player_names(url, selector)
        
        if not player_names:
            await ctx.send("No player names found on the webpage with the provided selector.")
            return
        
        # Create a text file with the player names
        output_file_name = "player_names.txt"
        with open(output_file_name, "w", encoding="utf-8") as f:
            for name in player_names:
                f.write(f"{name}\n")
        
        # Send the results
        await ctx.send(f"Found {len(player_names)} player names.", file=discord.File(output_file_name))
    
    except Exception as e:
        # Truncate error message if it's too long to avoid Discord's 2000 character limit
        error_msg = str(e)
        if len(error_msg) > 1500:
            error_msg = error_msg[:1500] + "... (error message truncated)"
        
        await ctx.send(f"Error extracting player names: {error_msg}")

# Specialized function for Lepoker.io player extraction
async def extract_lepoker_player_names(url, selector):
    async with async_playwright() as p:
        # Launch browser with optimized settings for Lepoker.io
        browser = await p.chromium.launch(
            headless=True,
            args=[
                '--disable-dev-shm-usage',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-web-security',
                '--disable-features=IsolateOrigins,site-per-process',
                '--disable-site-isolation-trials'
            ]
        )
        
        # Create context with larger viewport and mobile emulation (sometimes helps with lazy loading)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        
        # Create a new page with longer timeouts
        page = await context.new_page()
        page.set_default_timeout(180000)  # 3 minutes timeout
        
        try:
            print(f"Navigating to {url}")
            # Navigate to the URL with a longer timeout and wait until network is idle
            await page.goto(url, timeout=120000, wait_until="networkidle")
            
            # Wait for the page to be fully loaded
            await page.wait_for_load_state("networkidle", timeout=60000)
            await page.wait_for_timeout(5000)  # Additional 5 seconds wait
            
            # Check if we're on the correct page
            print("Page loaded, checking content")
            
            # Set to store unique player names
            player_names = set()
            
            # Function to extract names and return how many new names were found
            async def extract_names():
                elements = await page.locator(selector).all()
                initial_count = len(player_names)
                
                for element in elements:
                    try:
                        text = await element.text_content()
                        text = text.strip()
                        if text and len(text) > 0:
                            player_names.add(text)
                    except Exception:
                        continue
                
                return len(player_names) - initial_count
            
            # Initial extraction
            await extract_names()
            print(f"Initial extraction found {len(player_names)} players")
            
            # 1. Try small scrolls first
            print("Starting small scrolls...")
            for i in range(50):
                try:
                    # Scroll down a small amount
                    await page.evaluate(f"window.scrollBy(0, 300)")
                    await page.wait_for_timeout(500)  # Short wait
                    
                    # Every 5 scrolls, check for new names
                    if i % 5 == 0:
                        new_names = await extract_names()
                        print(f"Scroll group {i//5+1}: Found {new_names} new players, total: {len(player_names)}")
                        
                        # If we've found all 635 players or more, we can stop
                        if len(player_names) >= 635:
                            print(f"Found {len(player_names)} players, which is >= 635, stopping scrolling")
                            break
                except Exception as e:
                    print(f"Error during small scroll {i+1}: {str(e)}")
            
            # 2. If we don't have enough players yet, try larger jumps
            if len(player_names) < 635:
                print("Trying larger scroll jumps...")
                for i in range(20):
                    try:
                        # Larger scroll jump
                        await page.evaluate(f"window.scrollBy(0, 1000)")
                        await page.wait_for_timeout(1000)  # Longer wait
                        
                        new_names = await extract_names()
                        print(f"Large scroll {i+1}: Found {new_names} new players, total: {len(player_names)}")
                        
                        # If we've found all 635 players or more, we can stop
                        if len(player_names) >= 635:
                            print(f"Found {len(player_names)} players, which is >= 635, stopping scrolling")
                            break
                    except Exception as e:
                        print(f"Error during large scroll {i+1}: {str(e)}")
            
            # 3. If we still don't have enough, try scrolling to specific positions
            if len(player_names) < 635:
                print("Trying scrolling to specific positions...")
                scroll_positions = [0.2, 0.4, 0.6, 0.8, 1.0]  # Percentage of page height
                
                for pos in scroll_positions:
                    try:
                        # Scroll to percentage of page height
                        await page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pos})")
                        await page.wait_for_timeout(2000)  # Longer wait
                        
                        new_names = await extract_names()
                        print(f"Position scroll {pos*100}%: Found {new_names} new players, total: {len(player_names)}")
                    except Exception as e:
                        print(f"Error during position scroll {pos*100}%: {str(e)}")
            
            # 4. Try clicking pagination elements if they exist
            if len(player_names) < 635:
                print("Looking for pagination elements...")
                pagination_selectors = [
                    ".pagination button",
                    ".pagination a",
                    "button.pagination-next",
                    "a.pagination-next",
                    ".next-page",
                    "button:has-text('Next')",
                    "a:has-text('Next')"
                ]
                
                for pagination_selector in pagination_selectors:
                    try:
                        pagination_elements = await page.locator(pagination_selector).all()
                        if len(pagination_elements) > 0:
                            print(f"Found {len(pagination_elements)} pagination elements with selector: {pagination_selector}")
                            
                            # Click each pagination element
                            for i, element in enumerate(pagination_elements):
                                try:
                                    await element.click()
                                    await page.wait_for_timeout(2000)  # Wait for content to load
                                    
                                    new_names = await extract_names()
                                    print(f"Pagination click {i+1}: Found {new_names} new players, total: {len(player_names)}")
                                except Exception:
                                    continue
                            
                            break
                    except Exception:
                        continue
            
            # 5. Try a specialized approach for Lepoker.io - force load all players with JavaScript
            if len(player_names) < 635:
                print("Trying specialized JavaScript approach for Lepoker.io...")
                try:
                    # This JavaScript tries to trigger the loading of all players
                    await page.evaluate("""() => {
                        // Try to find and click any "load more" buttons
                        const buttons = Array.from(document.querySelectorAll('button'));
                        for (const button of buttons) {
                            if (button.textContent.toLowerCase().includes('more') || 
                                button.textContent.toLowerCase().includes('all') ||
                                button.textContent.toLowerCase().includes('show')) {
                                button.click();
                            }
                        }
                        
                        // Scroll to various positions to trigger lazy loading
                        const scrollPositions = [0.2, 0.4, 0.6, 0.8, 1.0];
                        for (const pos of scrollPositions) {
                            window.scrollTo(0, document.body.scrollHeight * pos);
                        }
                        
                        // Try to find the player container and force display
                        const containers = document.querySelectorAll('div');
                        for (const container of containers) {
                            if (container.querySelectorAll('div.truncate').length > 10) {
                                container.style.maxHeight = 'none';
                                container.style.overflow = 'visible';
                            }
                        }
                    }""")
                    
                    await page.wait_for_timeout(3000)  # Wait for changes to take effect
                    new_names = await extract_names()
                    print(f"JavaScript approach: Found {new_names} new players, total: {len(player_names)}")
                except Exception as e:
                    print(f"Error in JavaScript approach: {str(e)}")
            
            # 6. Final extraction using JavaScript to get all elements
            print("Performing final extraction with JavaScript...")
            try:
                # This JavaScript tries to extract all player names directly from the DOM
                all_texts = await page.evaluate(f"""() => {{
                    // Try the specified selector first
                    let elements = Array.from(document.querySelectorAll('{selector}'));
                    
                    // If that doesn't work well, try other common selectors
                    if (elements.length < 10) {{
                        const selectors = [
                            'div.truncate',
                            '.player-name',
                            '.player-username',
                            '.player-list div.truncate',
                            'tr td div.truncate',
                            'tr td:nth-child(2)',
                            '.players-table tr td:nth-child(2)',
                            '.players-list div.truncate'
                        ];
                        
                        for (const sel of selectors) {{
                            const newElements = Array.from(document.querySelectorAll(sel));
                            if (newElements.length > elements.length) {{
                                elements = newElements;
                            }}
                        }}
                    }}
                    
                    // Extract text content from elements
                    return elements
                        .map(el => el.textContent.trim())
                        .filter(text => text.length > 0);
                }}""")
                
                for text in all_texts:
                    player_names.add(text)
                
                print(f"Final JavaScript extraction found {len(player_names)} total players")
            except Exception as e:
                print(f"Error in final JavaScript extraction: {str(e)}")
            
            # 7. Try one more approach - get all text content from the page and parse it
            if len(player_names) < 635:
                print("Trying to extract from full page content...")
                try: 
                    full_page_text = await page.evaluate("""() => {
                        return document.body.innerText;
                    }""")
                    
                    # Split by newlines and process each line
                    lines = full_page_text.split('\n')
                    for line in lines:
                        line = line.strip()
                        if line and len(line) > 0 and len(line) < 50:  # Likely a username if not too long
                            player_names.add(line)
                    
                    print(f"Full page content extraction found {len(player_names)} total players")
                except Exception as e:
                    print(f"Error extracting from full page: {str(e)}")
            
            # Convert set back to list
            result = list(player_names)
            
            # Filter out non-player entries
            filtered_result = [
                name for name in result 
                if not name.isdigit() and 
                name not in ["Players", "Rank", "Clan", "Chips", "Tables", "Finished", "-", "Name", "Player", "Username"] and
                len(name) > 1  # Exclude single characters
            ]
            
            print(f"Final result: {len(filtered_result)} players after filtering")
            return filtered_result if filtered_result else result
        
        finally:
            await browser.close()

@bot.command(name="playhelp")
async def playhelp(ctx):
    if not is_play_authorized(ctx):
        await ctx.message.delete()
        return
    help_text = (
        "**Play Bot Help Menu - Detailed Guide**\n\n"
        "**Event Creation and Settings Management**\n"
        "With this bot, you can create various events (tournaments, competitions, etc.) in your server, "
        "set event-specific links, channels, interaction limits, same nickname limits, and maintain an Excel file.\n\n"
        "**Commands and Example Usage:**\n\n"
        "1. **!createplayevent <event_name>**\n"
        "   - Description: Creates a new event.\n"
        "   - Example: `!createplayevent Tournament2025`\n\n"
        "2. **!setplaylink <event_name> <link> [password xxxxx]**\n"
        "   - Description: Sets the link and optional password for the event.\n"
        "   - Example: `!setplaylink Tournament2025 https://example.com/tournament password 12345`\n\n"
        "3. **!setplaychannel <event_name> <channelID or #channel>**\n"
        "   - Description: Sets the channel where the event's button will be sent.\n\n"
        "4. **!setauthorizedrole <roleID or @role>**\n"
        "   - Description: Sets the authorized role for play event commands.\n\n"
        "5. **!setallowedrole <roleID1,roleID2,...>**\n"
        "   - Description: Specifies roles allowed to use the button.\n\n"
        "6. **!removeallowedrole <roleID1,roleID2,...>**\n"
        "   - Description: Removes the specified roles from the allowed list.\n\n"
        "7. **!samenicknamefilter <event_name> on <limit> / off**\n"
        "   - Description: Limits the number of times the same in-game username can be entered.\n\n"
        "8. **!sendplay <event_name> [channelID]**\n"
        "   - Description: Sends the Play button for the event.\n\n"
        "9. **!sendplaylimit <event_name> <roleID or @role> <limit>**\n"
        "   - Description: Sets the interaction limit for the specified role for the event.\n\n"
        "10. **!sendplaysettings <event_name>**\n"
        "    - Description: Displays all event settings.\n\n"
        "11. **!getplayexcel <event_name>**\n"
        "    - Description: Sends the Excel file.\n\n"
        "12. **!deletesendplay <event_name>**\n"
        "    - Description: Deletes the event and associated files (usage data is cleared).\n\n"
        "13. **!playlistid <event_name>**\n"
        "    - Description: Lists Discord IDs of event participants in a text file. The IDs are arranged side by side (separated by spaces), with every 150 IDs starting a new paragraph.\n\n"
        "14. **!removeplaybutton <event_name>**\n"
        "    - Description: Removes the Play button for the specified event.\n\n"
        "15. **!allplaylist**\n"
        "    - Description: Lists all created events.\n\n"
        "16. **!checkgameusername <event_name> <username1 username2 ...>**\n"
        "    - Description: Checks if the provided game usernames match with Discord IDs in the event records. Now supports case-insensitive and fuzzy matching.\n"
        "    - Example: `!checkgameusername Tournament2025 Player1 Player2 Player3`\n\n"
        "17. **!checkgameusername id <event_name> <username1 username2 ...>**\n"
        "    - Description: Same as above, but outputs only the Discord IDs in the text file.\n"
        "    - Example: `!checkgameusername id Tournament2025 Player1 Player2 Player3`\n\n"
        "18. **!checkgameusernameid <event_name> <username1 username2 ...>**\n"
        "    - Description: Alias for the above command, outputs only Discord IDs.\n"
        "    - Example: `!checkgameusernameid Tournament2025 Player1 Player2 Player3`\n\n"
        "19. **!getusername <url> <selector>**\n"
        "    - Description: Extracts player names from a webpage using the specified selector and saves them to a text file.\n"
        "    - Example: `!getusername https://app.lepoker.io/m/lj2Dxdy/players \"div.truncate\"`\n\n"
        "20. **!playhelp**\n"
        "    - Description: Shows this help menu.\n"
    )
    
    # Split message into chunks if it exceeds 2000 characters
    chunks = _chunk_text_message(help_text, limit=2000)
    for chunk in chunks:
        await ctx.send(chunk)

# ---------------- CAPTCHA Verification Functions ----------------

def _check_captcha_rate_limit(user_id: int) -> bool:
    """Check if user is within rate limit for captcha requests"""
    current_time = time.time()
    user_requests = captcha_rate_limits[user_id]
    
    # Remove old requests outside the window
    user_requests[:] = [req_time for req_time in user_requests if current_time - req_time < CAPTCHA_RATE_WINDOW]
    
    # Check if user has exceeded rate limit
    if len(user_requests) >= CAPTCHA_RATE_LIMIT:
        return False
    
    return True

def _add_captcha_rate_limit_request(user_id: int):
    """Add a request to the rate limit tracker"""
    current_time = time.time()
    captcha_rate_limits[user_id].append(current_time)


def _generate_captcha_code(length: int = 6) -> str:
    # Avoid ambiguous characters like 0/O and 1/I
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(random.choice(alphabet) for _ in range(length))


def _create_text_image(code: str) -> bytes:
    """Creates simple text image"""
    width, height = 300, 100
    # Light background
    bg_color = (245, 245, 245)
    text_color = (30, 30, 30)
    
    image = Image.new("RGB", (width, height), bg_color)
    draw = ImageDraw.Draw(image)
    
    # Font loading
    font = None
    possible_fonts = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for fp in possible_fonts:
        try:
            font = ImageFont.truetype(fp, 48)
            break
        except Exception:
            pass
    if font is None:
        font = ImageFont.load_default()
    
    # Center text
    try:
        bbox = draw.textbbox((0, 0), code, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except AttributeError:
        # Old PIL version
        text_width, text_height = draw.textsize(code, font=font)
    
    x = (width - text_width) // 2
    y = (height - text_height) // 2
    
    # Draw text
    draw.text((x, y), code, font=font, fill=text_color)
    
    # Add light noise
    for _ in range(50):
        x_noise = random.randint(0, width - 1)
        y_noise = random.randint(0, height - 1)
        image.putpixel((x_noise, y_noise), (
            random.randint(200, 240),
            random.randint(200, 240), 
            random.randint(200, 240)
        ))
    
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


class CaptchaModal(discord.ui.Modal):
    def __init__(self, expected_code: str, verify_role_id: int, user_id: int = None, show_code_hint: bool = False):
        super().__init__(title="Captcha Verification")
        self.expected_code = expected_code
        self.verify_role_id = verify_role_id
        self.user_id = user_id
        self.answer_input = discord.ui.TextInput(
            label=(f"Enter code: {expected_code}" if show_code_hint else "Enter code"),
            placeholder="Type the code here",
            min_length=1,
            max_length=12,
            required=True,
        )
        self.add_item(self.answer_input)

    async def on_submit(self, interaction: discord.Interaction):
        provided = (self.answer_input.value or "").strip()
        if provided != self.expected_code:
            await interaction.response.send_message(
                "Incorrect code. Please click the verification button again to retry.",
                ephemeral=True,
            )
            return

        guild = interaction.guild
        if guild is None:
            await interaction.response.send_message(
                "This action can only be performed within a server.", ephemeral=True
            )
            return

        role = guild.get_role(self.verify_role_id)
        if role is None:
            await interaction.response.send_message(
                "Verification role not found. Please contact an administrator.",
                ephemeral=True,
            )
            return

        # Ensure we have a Member object
        member = interaction.user
        if not isinstance(member, discord.Member):
            try:
                member = await guild.fetch_member(interaction.user.id)
            except Exception:
                member = None

        if member is None:
            await interaction.response.send_message(
                "Unable to retrieve member information.", ephemeral=True
            )
            return

        # If already verified
        if any(r.id == role.id for r in getattr(member, "roles", [])):
            await interaction.response.send_message(
                "You are already verified.", ephemeral=True
            )
            return

        try:
            await member.add_roles(role, reason="Captcha verified")
        except discord.Forbidden:
            await interaction.response.send_message(
                "Unable to assign role: Bot lacks permissions or role hierarchy prevents this action.",
                ephemeral=True,
            )
            return
        except Exception:
            await interaction.response.send_message(
                "An error occurred while assigning the role.", ephemeral=True
            )
            return

        await interaction.response.send_message(
            f"Success! {role.mention} role has been assigned.", ephemeral=True
        )
        
        # Clean up session after successful verification
        if self.user_id:
            active_captcha_sessions.discard(self.user_id)
            print(f"[captcha] Successfully verified user {self.user_id}")


class CaptchaVerifyView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(
        label="Verify",
        style=discord.ButtonStyle.success,
        custom_id="captcha_verify_button",
    )
    async def verify_button(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ):
        # This method is now handled by the on_interaction event handler
        # to prevent double acknowledgment issues
        pass


class CaptchaCodeEntryView(discord.ui.View):
    def __init__(self, expected_code: str, verify_role_id: int, user_id: int):
        super().__init__(timeout=180)
        self.expected_code = expected_code
        self.verify_role_id = verify_role_id
        self.user_id = user_id

    @discord.ui.button(label="Enter Code", style=discord.ButtonStyle.primary, custom_id="captcha_enter_code")
    async def enter_code(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ):
        # Only allow the original user to use this button
        if interaction.user.id != self.user_id:
            await interaction.response.send_message(
                "This captcha is not for you.", ephemeral=True
            )
            return
            
        await interaction.response.send_modal(CaptchaModal(self.expected_code, self.verify_role_id, self.user_id))

    async def on_timeout(self):
        # Clean up session when view times out
        active_captcha_sessions.discard(self.user_id)
        print(f"[captcha] Session timeout for user {self.user_id}")

# ---------------- Regex Moderation Commands ----------------

# Define or update a regex rule
@bot.command(name="regex")
async def define_regex(ctx, regexsettingsname: str, *, regexcommand: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    name_key = regexsettingsname.strip().lower()
    # Accept extended syntaxes: /pattern/flags or plain pattern with optional --flags i m s x ...
    pattern_text, flags_letters = _parse_pattern_and_flags(regexcommand)
    try:
        compiled = _compile_with_flags(pattern_text, flags_letters)
    except re.error as e:
        await ctx.send(f"Invalid regex: {e}")
        return
    guild_id = ctx.guild.id
    if guild_id not in regex_settings_by_guild:
        regex_settings_by_guild[guild_id] = {}
    settings = regex_settings_by_guild[guild_id].get(name_key, {"channels": set(), "exempt_users": set(), "exempt_roles": set()})
    settings["pattern"] = regexcommand
    settings["compiled"] = compiled
    regex_settings_by_guild[guild_id][name_key] = settings
    save_settings()
    if settings["channels"]:
        ch_mentions = ", ".join(f"<#{cid}>" for cid in settings["channels"])
        await ctx.send(
            f"Regex setting updated: `{regexsettingsname}`\n"
            f"Pattern: `{regexcommand}`\n"
            f"Engine: `{_REGEX_ENGINE_NAME}`  Flags: `{flags_letters or '-'}`\n"
            f"Applied channels: {ch_mentions}"
        )
    else:
        await ctx.send(
            f"Regex setting saved: `{regexsettingsname}`\n"
            f"Pattern: `{regexcommand}`\n"
            f"Engine: `{_REGEX_ENGINE_NAME}`  Flags: `{flags_letters or '-'}`\n"
            f"No channels assigned yet. Use `!setregexsettings {regexsettingsname} <channels>` to assign."
        )

# Assign channels to a regex rule
@bot.command(name="setregexsettings")
async def set_regex_settings(ctx, regexsettingsname: str, *, channels: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    guild_id = ctx.guild.id
    name_key = regexsettingsname.strip().lower()
    guild_rules = regex_settings_by_guild.get(guild_id)
    if not guild_rules or name_key not in guild_rules:
        await ctx.send("Please create the regex rule first: `!regex <regexsettingsname> <regex>`")
        return
    tokens = channels.replace(",", " ").split()
    lower_tokens = [t.lower() for t in tokens]
    selected: set[int] = set()
    invalid: list[str] = []

    # Helper: parse channel-like token into channel id (if valid)
    def _parse_channel_token(token: str):
        raw = token.strip()
        if raw.startswith("<#") and raw.endswith(">"):
            raw = raw[2:-1]
        try:
            cid_local = int(raw)
        except ValueError:
            return None
        channel_local = ctx.guild.get_channel(cid_local)
        if channel_local is None:
            return None
        return cid_local

    if "allchannel" in lower_tokens:
        # Start with all text channels
        for ch in ctx.guild.text_channels:
            selected.add(ch.id)

        if "notchannel" in lower_tokens:
            idx = lower_tokens.index("notchannel")
            exclude_tokens = tokens[idx + 1 :]
            if not exclude_tokens:
                # No exclusions provided; proceed with all text channels
                pass
            else:
                exclude_ids: set[int] = set()
                for tok in exclude_tokens:
                    cid = _parse_channel_token(tok)
                    if cid is None:
                        invalid.append(tok)
                    else:
                        exclude_ids.add(cid)
                selected -= exclude_ids
    else:
        if "notchannel" in lower_tokens:
            await ctx.send("Use `notchannel` only together with `allchannel`. Example: `!setregexsettings spamRule allchannel notchannel #channel1 #channel2`")
            return
        for tok in tokens:
            cid = _parse_channel_token(tok)
            if cid is None:
                invalid.append(tok)
                continue
            selected.add(cid)

    if not selected:
        await ctx.send("Please specify valid channels. Examples: `!setregexsettings spamRule #general #chat` or `!setregexsettings spamRule allchannel notchannel #log #mod`")
        return

    guild_rules[name_key]["channels"] = selected
    save_settings()
    ch_mentions = ", ".join(f"<#{cid}>" for cid in selected)
    msg = f"Applied channels updated for `{regexsettingsname}`: {ch_mentions}"
    if invalid:
        msg += f"\nIgnored/Invalid: {' '.join(invalid)}"
    await ctx.send(msg)

# Set exemptions (users or roles) for a regex rule
@bot.command(name="setregexexempt")
async def set_regex_exempt(ctx, regexsettingsname: str, kind: str, *, targets: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    guild_id = ctx.guild.id
    name_key = regexsettingsname.strip().lower()
    guild_rules = regex_settings_by_guild.get(guild_id)
    if not guild_rules or name_key not in guild_rules:
        await ctx.send("Please create the regex rule first: `!regex <regexsettingsname> <regex>`")
        return
    kind_l = kind.strip().lower()
    if kind_l not in ("users", "roles"):
        await ctx.send("Please specify a type: `users` or `roles`. Example: `!setregexexempt spam users @u1 @u2` or `!setregexexempt spam roles @role1 @role2`")
        return
    tokens = targets.replace(",", " ").split()
    selected: set[int] = set()
    invalid = []
    for tok in tokens:
        raw = tok.strip()
        # Normalize mentions
        if kind_l == "roles":
            if raw.startswith("<@&") and raw.endswith(">"):
                raw = raw[3:-1]
        else:  # users
            if raw.startswith("<@!") and raw.endswith(">"):
                raw = raw[3:-1]
            elif raw.startswith("<@") and raw.endswith(">"):
                raw = raw[2:-1]
        try:
            _id = int(raw)
        except ValueError:
            invalid.append(tok)
            continue
        if kind_l == "roles":
            role = ctx.guild.get_role(_id)
            if role is None:
                invalid.append(tok)
                continue
        else:
            member = ctx.guild.get_member(_id)
            if member is None:
                invalid.append(tok)
                continue
        selected.add(_id)
    if not selected:
        await ctx.send("Please specify valid targets. Examples:\n- `!setregexexempt spam users @alice @bob`\n- `!setregexexempt spam roles @Admin 123456789012345678`")
        return
    if kind_l == "roles":
        guild_rules[name_key]["exempt_roles"] = selected
        save_settings()
        names = []
        for rid in selected:
            role = ctx.guild.get_role(rid)
            if role:
                names.append(role.name)
            else:
                names.append(f"Unknown({rid})")
        names_str = ", ".join(names)
        msg = f"Exempt roles updated for `{regexsettingsname}`: {names_str}"
    else:
        guild_rules[name_key]["exempt_users"] = selected
        save_settings()
        names = []
        for uid in selected:
            member = ctx.guild.get_member(uid)
            if member:
                names.append(member.name)
            else:
                names.append(f"Unknown({uid})")
        names_str = ", ".join(names)
        msg = f"Exempt users updated for `{regexsettingsname}`: {names_str}"
    if invalid:
        msg += f"\nIgnored/Invalid: {' '.join(invalid)}"
    await ctx.send(msg)

# Show regex settings (all or specific)
@bot.command(name="regexsettings")
async def regexsettings(ctx, regexsettingsname: str = None):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    guild_id = ctx.guild.id
    guild_rules = regex_settings_by_guild.get(guild_id)
    if not guild_rules:
        await ctx.send("There are no regex settings defined in this server.")
        return

    def _mentions_list(ids: set[int], kind: str) -> str:
        """Convert a set of IDs to names/mentions. Channels use mentions, users/roles use names."""
        if not ids:
            return "None"
        ids_list = list(ids)
        if kind == "channel":
            return ", ".join(f"<#{i}>" for i in ids_list)
        elif kind == "user":
            names = []
            for uid in ids_list:
                member = ctx.guild.get_member(uid)
                if member:
                    names.append(member.name)
                else:
                    names.append(f"Unknown({uid})")
            return ", ".join(names)
        elif kind == "role":
            names = []
            for rid in ids_list:
                role = ctx.guild.get_role(rid)
                if role:
                    names.append(role.name)
                else:
                    names.append(f"Unknown({rid})")
            return ", ".join(names)
        return "None"

    if regexsettingsname:
        name_key = regexsettingsname.strip().lower()
        rule = guild_rules.get(name_key)
        if not rule:
            await ctx.send("No regex setting found with the specified name.")
            return
        pattern_text = rule.get("pattern", "-")
        
        channels = rule.get("channels", set())
        exempt_users = rule.get("exempt_users", set())
        exempt_roles = rule.get("exempt_roles", set())
        status = "Active" if channels else "Inactive"

        settings_text = f"**Regex Settings - {regexsettingsname}**\n\n"
        settings_text += f"**Status:** {status}\n"
        settings_text += f"**Pattern:** `{pattern_text}`\n"
        settings_text += f"**Applied Channels:** {_mentions_list(channels, 'channel')}\n"
        settings_text += f"**Exempt Users:** {_mentions_list(exempt_users, 'user')}\n"
        settings_text += f"**Exempt Roles:** {_mentions_list(exempt_roles, 'role')}\n"
        
        # Split message if it exceeds 2000 characters
        chunks = _chunk_text_message(settings_text, limit=2000)
        for chunk in chunks:
            await ctx.send(chunk)
        return

    # List all rules - Split into multiple messages if needed
    rules_list = list(guild_rules.items())
    if not rules_list:
        await ctx.send("No regex settings found in this server.")
        return
    
    # Build message text
    message_chunks = []
    current_message = "**Regex Settings**\n\n"
    
    for name_key, rule in rules_list:
        pattern_text = rule.get("pattern", "-")
        
        channels = rule.get("channels", set())
        exempt_users = rule.get("exempt_users", set())
        exempt_roles = rule.get("exempt_roles", set())
        status = "Active" if channels else "Inactive"
        
        # Get all mentions without limit
        channels_text = _mentions_list(channels, "channel")
        users_text = _mentions_list(exempt_users, "user")
        roles_text = _mentions_list(exempt_roles, "role")
        
        rule_text = (
            f"**{name_key}**\n"
            f"Status: {status}\n"
            f"Pattern: `{pattern_text}`\n"
            f"Channels: {channels_text}\n"
            f"Exempt Users: {users_text}\n"
            f"Exempt Roles: {roles_text}\n\n"
        )
        
        # Check if adding this rule would exceed message limit
        if len(current_message) + len(rule_text) > 1900:
            message_chunks.append(current_message)
            current_message = "**Regex Settings (Continued)**\n\n" + rule_text
        else:
            current_message += rule_text
    
    # Add the last message chunk
    if current_message:
        message_chunks.append(current_message)
    
    # Send all chunks
    for chunk in message_chunks:
        await ctx.send(chunk)

# Delete a regex setting by name
@bot.command(name="delregexsettings")
async def delregexsettings(ctx, regexsettingsname: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    guild_id = ctx.guild.id
    guild_rules = regex_settings_by_guild.get(guild_id)
    if not guild_rules:
        await ctx.send("There are no regex settings defined in this server.")
        return
    name_key = regexsettingsname.strip().lower()
    if name_key not in guild_rules:
        await ctx.send("No regex setting found with the specified name.")
        return
    del guild_rules[name_key]
    if not guild_rules:
        try:
            del regex_settings_by_guild[guild_id]
        except KeyError:
            pass
    save_settings()
    await ctx.send(f"Regex setting deleted: `{regexsettingsname}`")

# ---------------- CAPTCHA Verification Commands ----------------

@bot.event
async def on_ready():
    # Load settings from files
    load_settings()
    load_security_settings()
    load_spam_violation_stats()
    
    # Register persistent view so button keeps working after restart
    try:
        bot.add_view(CaptchaVerifyView())
    except Exception:
        pass
    print(f"Logged in as {bot.user} (ID: {getattr(bot.user, 'id', '-')})")
    print("[SETTINGS] Bot ready with loaded settings")


@bot.command(name="setverifyrole")
async def setverifyrole(ctx, role_identifier: str):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    global captcha_verify_role_id

    raw = role_identifier.strip()
    if raw.startswith("<@&") and raw.endswith(">"):
        raw = raw[3:-1]
    try:
        rid = int(raw)
    except ValueError:
        await ctx.send("Please enter a valid role ID or role mention.")
        return
    role = ctx.guild.get_role(rid)
    if role is None:
        await ctx.send("No role found with this ID.")
        return
    captcha_verify_role_id = rid
    save_settings()
    await ctx.send(f"Verification role set: {role.mention} ({rid})")


@bot.command(name="setverifypaneltext")
async def setverifypaneltext(ctx, text_type: str, *, content: str):
    # Debug: Command triggered
    print(f"[DEBUG] setverifypaneltext command triggered by {ctx.author} with type: {text_type}")
    
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        print(f"[DEBUG] User {ctx.author} not authorized")
        return
    
    print(f"[DEBUG] User authorized, processing...")
    
    guild_id = ctx.guild.id
    text_type = text_type.lower().strip()
    
    print(f"[DEBUG] Text type: {text_type}, Content length: {len(content)}")
    
    if text_type not in ["title", "description", "image"]:
        await ctx.send("Please specify either `title`, `description`, or `image`. Examples:\n• `!setverifypaneltext title Welcome to Our Server`\n• `!setverifypaneltext image https://example.com/image.png`")
        print(f"[DEBUG] Invalid text type: {text_type}")
        return
    
    if text_type == "image":
        print(f"[DEBUG] Processing image URL: {content[:100]}...")
        
        # Validate URL format (expanded check for various platforms)
        content_lower = content.lower()
        
        # Check for valid URL protocols
        valid_protocols = [
            "http://", "https://", "blob:", "data:"
        ]
        
        print(f"[DEBUG] Checking protocols...")
        if not any(content.startswith(protocol) for protocol in valid_protocols):
            await ctx.send("Image must be a valid URL starting with http://, https://, blob:, or data:")
            print(f"[DEBUG] Invalid protocol in URL: {content[:50]}")
            return
        
        print(f"[DEBUG] Protocol check passed")
        
        # Special handling for different URL types
        print(f"[DEBUG] Starting platform validation...")
        is_discord_cdn = ("cdn.discordapp.com" in content_lower or 
                         "media.discordapp.net" in content_lower or
                         "images-ext-1.discordapp.net" in content_lower or
                         "images-ext-2.discordapp.net" in content_lower or
                         "discordapp.com/attachments" in content_lower or
                         "discord.com/attachments" in content_lower)
        print(f"[DEBUG] Discord CDN check: {is_discord_cdn}")
        is_whatsapp = "web.whatsapp.com" in content_lower or "whatsapp" in content_lower
        is_blob_url = content.startswith("blob:")
        is_data_url = content.startswith("data:image/")
        is_gif_platform = any(platform in content_lower for platform in [
            "giphy.com", "tenor.com", "gfycat.com", "reddit.com", "redgifs.com"
        ])
        is_video_platform = any(platform in content_lower for platform in [
            "youtube.com", "youtu.be", "vimeo.com", "streamable.com", "twitch.tv", "tiktok.com"
        ])
        is_special_platform = any(platform in content_lower for platform in [
            "imgur.com", "gyazo.com", "prntscr.com", "lightshot.com", 
            "github.com", "githubusercontent.com", "telegram.org", "steamcommunity.com"
        ])
        
        # Check file extensions for regular URLs (images and videos)
        valid_extensions = [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg", ".mp4", ".mov", ".avi", ".webm", ".mkv"]
        has_valid_extension = any(content_lower.endswith(ext) for ext in valid_extensions)
        
        # Debug: Show which validations passed/failed
        print(f"[DEBUG] All platform checks completed")
        print(f"[DEBUG] Final validation results:")
        print(f"[DEBUG] - Discord CDN: {is_discord_cdn}")
        print(f"[DEBUG] - WhatsApp: {is_whatsapp}")
        print(f"[DEBUG] - Blob URL: {is_blob_url}")
        print(f"[DEBUG] - Data URL: {is_data_url}")
        print(f"[DEBUG] - GIF Platform: {is_gif_platform}")
        print(f"[DEBUG] - Video Platform: {is_video_platform}")
        print(f"[DEBUG] - Special Platform: {is_special_platform}")
        print(f"[DEBUG] - Valid Extension: {has_valid_extension}")
        
        debug_info = f"🔍 **URL Validation Debug:**\n"
        debug_info += f"URL: `{content[:100]}{'...' if len(content) > 100 else ''}`\n"
        debug_info += f"Discord CDN: {'✅' if is_discord_cdn else '❌'}\n"
        debug_info += f"WhatsApp: {'✅' if is_whatsapp else '❌'}\n"
        debug_info += f"Blob URL: {'✅' if is_blob_url else '❌'}\n"
        debug_info += f"Data URL: {'✅' if is_data_url else '❌'}\n"
        debug_info += f"GIF Platform: {'✅' if is_gif_platform else '❌'}\n"
        debug_info += f"Video Platform: {'✅' if is_video_platform else '❌'}\n"
        debug_info += f"Special Platform: {'✅' if is_special_platform else '❌'}\n"
        debug_info += f"Valid Extension: {'✅' if has_valid_extension else '❌'}\n"
        
        # Allow URL if it meets any of these criteria
        validation_passed = (is_discord_cdn or is_whatsapp or is_blob_url or is_data_url or 
                            is_gif_platform or is_video_platform or is_special_platform or has_valid_extension)
        print(f"[DEBUG] Overall validation result: {validation_passed}")
        
        if not validation_passed:
            message_text = (
                "❌ **Image URL Validation Failed**\n\n"
                f"{debug_info}\n\n"
                "**Supported URL Types:**\n"
                "• End with a valid image/video extension (.png, .jpg, .jpeg, .gif, .webp, .bmp, .svg, .mp4, .mov, .avi, .webm, .mkv)\n"
                "• Be a Discord CDN link (cdn.discordapp.com, discord.com/attachments)\n"
                "• Be a WhatsApp Web link\n"
                "• Be a blob: or data: URL\n"
                "• Be from a GIF platform (Giphy, Tenor, Gfycat, Reddit)\n"
                "• Be from a video platform (YouTube, Vimeo, Streamable, TikTok)\n"
                "• Be from a supported platform (Imgur, GitHub, Steam, etc.)"
            )
            await ctx.send(message_text)
            return
        else:
            # Success - show which validation passed
            print(f"[DEBUG] Validation passed! Sending success message...")
            success_message = f"✅ **Image URL Validation Passed**\n\n{debug_info}"
            await ctx.send(success_message)
            print(f"[DEBUG] Success message sent")
    
    if len(content) > 256 and text_type == "title":
        await ctx.send("Title must be 256 characters or less.")
        return
    
    if len(content) > 2048 and text_type == "description":
        await ctx.send("Description must be 2048 characters or less.")
        return
    
    if guild_id not in captcha_panel_texts:
        captcha_panel_texts[guild_id] = {
            "title": DEFAULT_PANEL_TITLE,
            "description": DEFAULT_PANEL_DESCRIPTION,
            "image": None
        }
    
    captcha_panel_texts[guild_id][text_type] = content
    save_settings()
    
    if text_type == "image":
        await ctx.send(f"Verification panel image updated successfully!\nImage URL: {content}")
    else:
        await ctx.send(f"Verification panel {text_type} updated successfully!")


@bot.command(name="showverifypaneltext")
async def showverifypaneltext(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    guild_id = ctx.guild.id
    panel_text = captcha_panel_texts.get(guild_id, {
        "title": DEFAULT_PANEL_TITLE,
        "description": DEFAULT_PANEL_DESCRIPTION,
        "image": None
    })
    
    message_text = "**Current Verification Panel Text**\n\n"
    message_text += f"**Title:**\n```{panel_text['title']}```\n\n"
    message_text += f"**Description:**\n```{panel_text['description']}```\n\n"
    
    image_url = panel_text.get('image')
    if image_url:
        message_text += f"**Image URL:**\n```{image_url}```\n\n"
    else:
        message_text += "**Image URL:**\n```Not set```\n\n"
    
    message_text += (
        "**Usage:**\n"
        "• `!setverifypaneltext title <new title>`\n"
        "• `!setverifypaneltext description <new description>`\n"
        "• `!setverifypaneltext image <image_url>`"
    )
    
    await ctx.send(message_text)


@bot.command(name="resetverifypaneltext")
async def resetverifypaneltext(ctx):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return
    
    guild_id = ctx.guild.id
    if guild_id in captcha_panel_texts:
        del captcha_panel_texts[guild_id]
    save_settings()
    
    await ctx.send("Verification panel text reset to default values.")


@bot.command(name="sendverifypanel")
async def sendverifypanel(ctx, channel: str = None):
    if not is_security_authorized(ctx):
        await ctx.message.delete()
        return

    target_channel = ctx.channel
    if channel:
        raw = channel.strip()
        if raw.startswith("<#") and raw.endswith(">"):
            raw = raw[2:-1]
        try:
            cid = int(raw)
            ch = ctx.guild.get_channel(cid)
            if ch is not None:
                target_channel = ch
        except ValueError:
            pass

    # Get custom panel text for this guild or use defaults
    guild_id = ctx.guild.id
    panel_text = captcha_panel_texts.get(guild_id, {
        "title": DEFAULT_PANEL_TITLE,
        "description": DEFAULT_PANEL_DESCRIPTION,
        "image": None
    })

    # Create message with button
    message_text = f"**{panel_text['title']}**\n\n{panel_text['description']}"
    
    # If there's an image URL, send it separately since we can't embed without embeds
    image_url = panel_text.get("image")
    
    try:
        # Send the verification message with button
        await target_channel.send(content=message_text, view=CaptchaVerifyView())
        
        # If there's an image, send it as a separate message
        if image_url:
            await target_channel.send(image_url)
        
        await ctx.send(f"Verification panel sent to: {target_channel.mention}")
    except Exception as e:
        await ctx.send("Failed to send verification panel.")

# Get bot token from environment variable
bot_token = os.getenv("PLAYBOT")
if not bot_token:
    bot_token = "BOTTOKENHERE"  # Fallback to hardcoded token if env var not set

bot.run(bot_token)
